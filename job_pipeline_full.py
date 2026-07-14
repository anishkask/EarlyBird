#!/usr/bin/env python3
"""
================================================================
  EarlyBird - Job Search Pipeline
  Strategy: surface fresh PAID roles before the applicant flood.
================================================================

Aggregates roles from stable, terms-respecting endpoints (Greenhouse, Lever,
and Ashby ATS JSON APIs for a config-driven company watchlist, plus the
RemoteOK JSON board), filters to relevant paid roles in the US/remote, ranks
by freshness + profile keyword match + role type, dedupes across sources, and
writes a color-coded Excel tracker. For the top-ranked roles it researches a
likely outreach contact and drafts a short tailored message.

Sources, target companies, keywords, and filters live in config.py.
Secrets (API keys, identity) live in .env.

Run modes:
  python job_pipeline_full.py                 standard run (config.DEFAULT_HOURS)
  python job_pipeline_full.py --hours 24      custom lookback window
  python job_pipeline_full.py --fresh         fast poll: short window, no outreach
  python job_pipeline_full.py --scrape-only   scrape + Excel only, no Claude calls
"""

import os
import sys
import re
import json
import html
import time
import argparse
import warnings
from datetime import datetime, timezone, timedelta
from pathlib import Path
from urllib.parse import quote_plus

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

import config

warnings.filterwarnings("ignore")
load_dotenv()

# --- Identity (from .env only; never hard-coded) ---
YOUR_NAME     = os.getenv("YOUR_NAME", "")
YOUR_EMAIL    = os.getenv("YOUR_EMAIL", "")
YOUR_LINKEDIN = os.getenv("YOUR_LINKEDIN", "")
MY_BACKGROUND = os.getenv("MY_BACKGROUND", "")

# Windows consoles default to cp1252 and choke on non-ASCII; force UTF-8 so a
# stray character can never crash a run. Prints themselves stay ASCII.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

# claude-sonnet-4-20250514 was retired by Anthropic (API returns 404).
# Sonnet 5 is its documented drop-in replacement at the same price tier.
OUTREACH_MODEL = "claude-sonnet-5"
WEB_SEARCH_TOOL = {"type": "web_search_20250305", "name": "web_search", "max_uses": 3}

# ===========================================================================
# Classification and ranking
# ===========================================================================

_ALL_KEYWORDS = (
    config.PROFILE_KEYWORDS["high"]
    + config.PROFILE_KEYWORDS["core"]
    + config.PROFILE_KEYWORDS["bonus"]
)

# Cheap title-level pre-filter so scrapers skip obviously irrelevant roles
# (sales, marketing, recruiting) before parsing descriptions.
_TITLE_PREFILTER = [
    "engineer", "developer", "software", "ml engineer", "machine learning",
    " ai ", "ai/", "ai engineer", "full stack", "full-stack", "fullstack",
    "backend", "back-end", "frontend", "front-end", "programmer", "llm",
]

_NON_US = [
    "dublin", "luxembourg", "bengaluru", "bangalore", "hyderabad", "warsaw",
    "toronto", "vancouver", "canada", "india", "brazil", "ireland", "germany",
    "united kingdom", "london", "france", "paris", "europe", "singapore",
    "australia", "china", "japan", "hong kong", "mexico", "spain", "netherlands",
    "amsterdam", "sweden", "switzerland", "zurich", "italy", "israel", "tel aviv",
    "uae", "dubai", "new zealand", "south africa", "argentina", "colombia",
    "poland", "romania", "portugal", "lisbon", "ukraine", "korea", "taiwan",
    "philippines", "indonesia", "vietnam", "emea", "apac", "latam",
]
# Matched against _norm_loc() output; " india " is word-bounded so it cannot
# match "Indianapolis, Indiana". Multi-word names match the normalized text.
_NON_US = [x if x != "india" else " india " for x in _NON_US]

_DEGREE_RE = re.compile(
    r"(bachelor'?s?\s+degree\s+(is\s+)?required"
    r"|require[sd]?\s+(a\s+)?(bachelor|master|b\.?s\.?|b\.?a\.?|m\.?s\.?)\b"
    r"|degree\s+(is\s+)?required"
    r"|must\s+have\s+(a\s+)?(bachelor|master|degree)"
    r"|\b(bs|ba|ms)\s+in\s+(computer|engineering))",
    re.I,
)


def hours_ago(dt):
    """Hours elapsed since datetime dt. Returns 999 if dt is None/invalid."""
    if dt is None:
        return 999
    if not hasattr(dt, "astimezone"):
        return 999
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    delta = datetime.now(timezone.utc) - dt
    return int(delta.total_seconds() / 3600)


def _title_maybe_relevant(title):
    """Fast title-only gate used inside scrapers to bound description parsing."""
    if not title:
        return False
    t = title.lower()
    return any(k in t for k in _TITLE_PREFILTER)


def is_relevant(title, desc="", extra_keywords=None):
    """A role is relevant if any profile keyword appears in title or description."""
    text = (title + " " + desc).lower()
    if extra_keywords and any(k in text for k in extra_keywords):
        return True
    return any(k in text for k in _ALL_KEYWORDS)


def _norm_loc(loc):
    """Lowercase, punctuation collapsed to spaces, space-padded -- so short
    tokens can be word-bounded ("india" must not match "Indianapolis, Indiana")."""
    return " " + re.sub(r"[^a-z0-9]+", " ", (loc or "").lower()).strip() + " "


def is_usa_location(loc):
    """Keep US, remote, and unknown locations; drop clearly non-US ones."""
    if not loc or not isinstance(loc, str):
        return True  # unknown location -> keep (better recall)
    s = _norm_loc(loc)
    if any(x in s for x in _NON_US):
        return False
    return True


# Locations that count as remote-friendly and always pass a target-location
# filter. Deliberately excludes bare "us" (substring-matches "Austin").
_REMOTE_TOKENS = ["remote", "anywhere", "united states", "usa", "work from home", "wfh", "nationwide"]

# Remote markers and region-restriction parsing for location_accept().
_REMOTE_MARKERS = ["remote", "work from home", "wfh", "distributed", "anywhere"]
_REMOTE_REGION_RE = re.compile(r"remote\s*\(([^)]*)\)")


def _remote_us_eligible(s):
    """True if the (lowercase) location is remote AND US-eligible.
    "Remote (IN)" / "Remote (DE)" style restrictions to non-US regions fail;
    unrestricted "Remote" or any restriction group mentioning the US passes."""
    if not any(m in s for m in _REMOTE_MARKERS):
        return False
    groups = _REMOTE_REGION_RE.findall(s)
    if not groups:
        return True  # unrestricted remote
    for g in groups:
        tokens = re.split(r"[^a-z0-9]+", g)
        if ("us" in tokens or "usa" in tokens or "united states" in g
                or "worldwide" in g or "anywhere" in g or "americas" in g
                or "north america" in g or "global" in g):
            return True
    return False


def location_accept(loc):
    """Config-driven geography gate. Returns "remote", "metro", "unknown"
    (blank location, kept for recall), or None (drop).
    Remote is the top preference and gets a rank boost downstream; onsite or
    hybrid is accepted only in config.ACCEPTED_METROS."""
    if not loc or not isinstance(loc, str) or not loc.strip():
        return "unknown"
    s = loc.lower()
    if config.ACCEPT_REMOTE and _remote_us_eligible(s):
        return "remote"
    n = _norm_loc(loc)
    if any(f" {m} " in n for m in config.ACCEPTED_METROS):
        return "metro"
    return None


def matches_target_locations(loc, targets):
    """True if loc is blank/remote or contains one of the user's target locations."""
    if not targets:
        return True
    if not loc or not isinstance(loc, str):
        return True  # blank location -> keep (better recall)
    s = loc.lower()
    if any(t in s for t in _REMOTE_TOKENS):
        return True
    return any(t.lower() in s for t in targets if t)


# Word-bounded so "internal tools" / "international team" can never classify
# a role as an internship -- the old substring check silently dropped real
# full-time roles on every prior run.
_INTERN_RE = re.compile(r"\b(interns?|internships?|co-?op|co op)\b")


def role_type(title, desc=""):
    """Classify employment type: intern, apprenticeship (incl. fellowship /
    returnship), contract-to-hire, contract (incl. freelance/temp/1099/W2),
    part-time, or full-time."""
    full = (title + " " + desc).lower()
    t = title.lower()
    if _INTERN_RE.search(full):
        return "intern"
    if "apprentice" in t or "fellowship" in t or "returnship" in t:
        return "apprenticeship"
    if "contract to hire" in t or "contract-to-hire" in t or "c2h" in t or "contract-to-perm" in t:
        return "contract-to-hire"
    if (any(k in t for k in ["contract", "contractor", "freelance", "1099", "w-2"])
            or re.search(r"\b(temp|temporary|w2)\b", t)):
        return "contract"
    if "part-time" in t or "part time" in t:
        return "part-time"
    return "full-time"


# --- Internship guard (config.INCLUDE_INTERNSHIPS gates the whole family) ---
# Candidate is NOT an enrolled student until Fall 2027, so:
#   - explicit unpaid / for-credit postings are dropped;
#   - explicit current-enrollment requirements are dropped;
#   - everything ambiguous is KEPT and flagged in the Notes column -- this
#     guard must never become a silent-suppression bug.
_UNPAID_MARKERS = [
    "unpaid", "for credit", "academic credit", "course credit",
    "college credit", "credit only",
]
_ENROLLMENT_MARKERS = [
    "currently enrolled", "must be enrolled", "enrolled in a degree",
    "enrolled in an accredited", "enrolled at an accredited",
    "must be a current student", "current student status",
    "currently pursuing a bachelor", "currently pursuing a degree",
    "currently pursuing an undergraduate", "currently pursuing a master",
    "actively pursuing a degree", "must be pursuing a degree",
    "receive academic credit", "eligible to receive credit",
    "returning to school in", "returning to a degree program",
]
_PAID_MARKERS = ["$", "/hr", "/hour", "per hour", "hourly", "stipend",
                 "paid internship", "compensation", "salary"]
_OPEN_MARKERS = ["open to non-students", "not required to be enrolled",
                 "no enrollment requirement", "recent grad", "recent graduates",
                 "returnship"]


def internship_guard(title, desc):
    """For intern-classified roles: (keep, type_label, note).
    Drops only on explicit unpaid/credit or enrollment-required language;
    ambiguity keeps the role with a verification note."""
    d = f"{title} {desc}".lower()
    if any(k in d for k in _UNPAID_MARKERS):
        return False, "", ""
    if any(k in d for k in _ENROLLMENT_MARKERS):
        return False, "", ""
    label = "internship-paid" if any(k in d for k in _PAID_MARKERS) else "internship"
    note = "" if any(k in d for k in _OPEN_MARKERS) else "verify enrollment requirement"
    if label == "internship" and note:
        note = "verify pay & enrollment requirement"
    return True, label, note


def requires_degree(title, desc=""):
    """Best-effort detection of a conferred-degree requirement."""
    return bool(_DEGREE_RE.search(title + " " + desc))


_YEARS_EXP_RE = re.compile(
    r"(\d{1,2})\s*\+?\s*years?(?:\s+of)?\s+(?:[a-z\-]+\s+){0,3}experience", re.I
)

_PLAIN_IC_TITLES = [
    "software engineer", "full stack", "full-stack", "fullstack", "frontend",
    "front end", "front-end", "backend", "back end", "back-end", "ai engineer",
    "ml engineer", "machine learning engineer", "software developer", "web developer",
]


def years_required(text):
    """Highest years-of-experience figure mentioned in the text, or 0."""
    yrs = [int(m) for m in _YEARS_EXP_RE.findall(text or "")]
    return max(yrs) if yrs else 0


def _norm_title(title):
    """Lowercase, punctuation collapsed to spaces, space-padded -- so short
    exclude tokens like " lead " or " sre " are word-bounded and cannot match
    inside longer words ("leadership") or across punctuation ("Lead, Backend")."""
    return " " + re.sub(r"[^a-z0-9]+", " ", (title or "").lower()).strip() + " "


def is_excluded_seniority(title):
    """Titles above the entry/mid ceiling (config.SENIORITY_EXCLUDE) are dropped."""
    t = (title or "").lower()
    # "Member of Technical Staff" is an entry-possible AI-startup title; its
    # "staff" is not a seniority marker. "Senior Member of Technical Staff"
    # still matches on the remaining "senior". Likewise "Fellowship" is a
    # program name, not the "Distinguished Fellow" seniority title.
    t = t.replace("member of technical staff", "")
    t = t.replace("fellowship", "")
    return any(k in _norm_title(t) for k in config.SENIORITY_EXCLUDE)


def is_excluded_function(title):
    """Off-target functions (config.FUNCTION_EXCLUDE): security, IT, QA, etc."""
    return any(k in _norm_title(title) for k in config.FUNCTION_EXCLUDE)


def is_excluded_ops(title):
    """Ops/infra/database-ops roles are off-track (config.OPS_INFRA_EXCLUDE).
    Title is padded with spaces so word-bounded tokens like " dba " work."""
    t = f" {(title or '').lower()} "
    return any(k in t for k in config.OPS_INFRA_EXCLUDE)


def title_allowed(title):
    """True if the title matches an AI-startup pattern (TRACK_TITLE_ALLOW).
    These pass the track gate AND count as relevant even with a sparse
    description (e.g. "Member of Technical Staff" with no keyword overlap)."""
    t = f" {(title or '').lower()} "
    return any(a in t for a in config.TRACK_TITLE_ALLOW)


def title_on_track(title, extra_keywords=None):
    """The title itself must carry a track signal (config.TRACK_TITLE_KEYWORDS)
    or match an AI-startup pattern (config.TRACK_TITLE_ALLOW). Description
    keywords still boost rank but cannot rescue an off-track title."""
    if title_allowed(title):
        return True
    t = f" {(title or '').lower()} "
    if any(k in t for k in config.TRACK_TITLE_KEYWORDS):
        return True
    if extra_keywords and any(k in t for k in extra_keywords):
        return True
    return False


def is_excluded_company(company):
    """Mega-caps and excluded firms (config.EXCLUDE_COMPANIES) are dropped."""
    c = (company or "").lower()
    return any(k in c for k in config.EXCLUDE_COMPANIES)


def is_entry_mid(title):
    """Explicit entry/mid title signal (config.ENTRY_MID_SIGNALS); rank bonus."""
    t = (title or "").lower()
    return any(k in t for k in config.ENTRY_MID_SIGNALS)


def rank_score(j, extra_keywords=None):
    """
    Attainability-aware ranking. Rewards entry/mid level + smaller company +
    stack match + freshness; penalizes residual seniority and large boards.
    A realistic small-company role outranks a fresh staff role at a unicorn.
    """
    title = j.get("title", "")
    text = (title + " " + j.get("description", "")).lower()
    s = 0.0

    # Stack / profile keyword match
    s += sum(3 for kw in config.PROFILE_KEYWORDS["high"] if kw in text)
    s += sum(2 for kw in config.PROFILE_KEYWORDS["core"] if kw in text)
    s += sum(1 for kw in config.PROFILE_KEYWORDS["bonus"] if kw in text)
    if extra_keywords:
        s += sum(2 for kw in extra_keywords if kw in text)

    # Freshness (reduced weight so it does not dominate attainability)
    ha = j.get("hours_ago", 999)
    if ha < 1:
        s += 6
    elif ha < 6:
        s += 4
    elif ha < 24:
        s += 2
    elif ha < 48:
        s += 1

    # Attainability: entry/mid level signal
    if is_entry_mid(title):
        s += 5
    elif any(k in title.lower() for k in _PLAIN_IC_TITLES):
        s += 3

    # Company-size proxy: reward small boards, penalize large ones
    bs = j.get("board_size")
    if isinstance(bs, int):
        if bs <= config.SMALL_BOARD_THRESHOLD:
            s += 5
        elif bs > config.LARGE_BOARD_THRESHOLD:
            s -= 5

    # Employment type: contract family leads (income speed), then full-time.
    rt = j.get("role_type", "full-time")
    if rt in ("contract", "contract-to-hire"):
        s += 4
    elif rt == "full-time":
        s += 2
    elif rt in ("part-time", "apprenticeship"):
        s += 1

    # Remote is the top geographic preference
    if j.get("work_mode") == "remote":
        s += 4

    # Penalties
    if j.get("requires_degree") and config.DEGREE_REQUIREMENT_MODE == "deprioritize":
        s -= 4
    if is_excluded_seniority(title):  # should be filtered already; safety net
        s -= 10
    return round(s, 1)


def fresh_flag(ha):
    """Human urgency flag for the tracker."""
    if ha < 1:
        return "<1h!"
    if ha < 6:
        return "<6h"
    return ""


def job(title, company, location, url, dt, source, desc="", board_size=None):
    """Normalize one posting into the pipeline's common job dict."""
    ha = hours_ago(dt)
    # 1500 chars so the internship guard can see enrollment/pay language,
    # which usually sits deep in the posting.
    return {
        "title": title, "company": company, "location": location,
        "job_url": url, "hours_ago": ha, "source": source,
        "description": (desc or "")[:1500], "board_size": board_size,
    }


# ===========================================================================
# Watchlist (config-driven, with optional cached discovery)
# ===========================================================================

def _ats_slugs_from_name(name):
    """Generate candidate ATS slugs from a company name."""
    base = re.sub(r"[^a-z0-9\s-]", "", name.lower()).strip()
    hyphen = re.sub(r"\s+", "-", base)
    nospace = re.sub(r"\s+", "", base)
    first = base.split()[0] if base.split() else base
    return list(dict.fromkeys([s for s in (hyphen, nospace, first) if s]))


def _check_greenhouse(slug):
    try:
        r = requests.get(
            f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs",
            headers=HEADERS, timeout=3,
        )
        return r.status_code == 200 and bool(r.json().get("jobs"))
    except requests.RequestException:
        return False


def _check_lever(slug):
    try:
        r = requests.get(
            f"https://api.lever.co/v0/postings/{slug}?mode=json",
            headers=HEADERS, timeout=3,
        )
        return r.status_code == 200 and isinstance(r.json(), list) and bool(r.json())
    except requests.RequestException:
        return False


def _load_discovered():
    """Read cached, previously-verified discovered companies. No network."""
    path = Path(config.WATCHLIST_CACHE_FILE)
    if not path.exists():
        return []
    try:
        return json.load(open(path, encoding="utf-8")).get("companies", [])
    except Exception:
        return []


def build_watchlist():
    """Company watchlist = config.WATCHLIST plus any cached discoveries."""
    companies = [dict(c) for c in config.WATCHLIST]
    seen = {(c["ats_type"], c["slug"]) for c in companies}
    for c in _load_discovered():
        key = (c.get("ats_type"), c.get("slug"))
        if c.get("slug") and key not in seen:
            seen.add(key)
            companies.append(c)
    return companies


def refresh_watchlist_cache():
    """
    Verify companies from the cold-outreach cache against Greenhouse/Lever and
    persist the confirmed ones for future runs. Runs at the end of a full run;
    failures never affect the current results.
    """
    cold_cache = Path("data") / "cold_outreach_cache.json"
    if not cold_cache.exists():
        return
    try:
        names = [
            c.get("name", "")
            for c in json.load(open(cold_cache, encoding="utf-8")).get("companies", [])
            if c.get("name")
        ]
    except Exception as e:
        print(f"WARNING: could not read cold-outreach cache: {e}")
        return

    existing = {(c["ats_type"], c["slug"]) for c in build_watchlist()}
    verified = _load_discovered()
    verified_keys = {(c.get("ats_type"), c.get("slug")) for c in verified}
    added = 0
    for name in names[:60]:  # bound the work
        for slug in _ats_slugs_from_name(name)[:2]:
            if _check_greenhouse(slug) and ("greenhouse", slug) not in existing:
                verified.append({"name": name, "ats_type": "greenhouse", "slug": slug})
                verified_keys.add(("greenhouse", slug))
                added += 1
                break
            if _check_lever(slug) and ("lever", slug) not in existing:
                verified.append({"name": name, "ats_type": "lever", "slug": slug})
                verified_keys.add(("lever", slug))
                added += 1
                break
    try:
        json.dump(
            {"companies": verified, "built_at": datetime.now().isoformat()},
            open(config.WATCHLIST_CACHE_FILE, "w", encoding="utf-8"),
            indent=2,
        )
        if added:
            print(f"Watchlist cache updated: {added} new verified company(ies) for next run")
    except Exception as e:
        print(f"WARNING: could not write watchlist cache: {e}")


COMPANIES = build_watchlist()
GREENHOUSE_SLUGS = [(c["slug"], c["name"]) for c in COMPANIES if c.get("ats_type") == "greenhouse"]
LEVER_SLUGS = [(c["slug"], c["name"]) for c in COMPANIES if c.get("ats_type") == "lever"]
ASHBY_SLUGS = list(config.ASHBY_COMPANIES)


# ===========================================================================
# Sources (stable ATS JSON APIs + RemoteOK; scrapers do not filter by role
# type -- central processing handles relevance/paid/location/ranking)
# ===========================================================================

def scrape_greenhouse(max_h):
    """Greenhouse public board API, one request per watchlist company.
    Undated postings fall back to a 72h-old timestamp rather than being lost."""
    out = []
    for slug, name in GREENHOUSE_SLUGS:
        try:
            r = requests.get(
                f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true",
                headers=HEADERS, timeout=8,
            )
            if r.status_code != 200:
                continue
            postings = r.json().get("jobs", [])
            board_size = len(postings)
            for j in postings:
                title = j.get("title", "")
                if not _title_maybe_relevant(title):
                    continue
                dt = None
                for field in ("updated_at", "created_at"):
                    ts = j.get(field, "")
                    if ts:
                        try:
                            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                            break
                        except ValueError:
                            continue
                if dt is None:
                    dt = datetime.now(timezone.utc) - timedelta(hours=72)
                if hours_ago(dt) > max_h:
                    continue
                loc = j.get("location", {})
                loc_name = loc.get("name", "") if isinstance(loc, dict) else str(loc)
                desc = BeautifulSoup(j.get("content", ""), "html.parser").get_text()
                out.append(job(title, name, loc_name, j.get("absolute_url", ""), dt,
                               "Greenhouse", desc, board_size=board_size))
        except Exception as e:
            print(f"WARNING: [Greenhouse:{slug}] {e}")
    return out


def scrape_lever(max_h):
    """Lever public postings API, one request per watchlist company.
    Undated postings fall back to a 72h-old timestamp rather than being lost."""
    out = []
    for slug, name in LEVER_SLUGS:
        try:
            r = requests.get(
                f"https://api.lever.co/v0/postings/{slug}?mode=json",
                headers=HEADERS, timeout=8,
            )
            if r.status_code != 200:
                continue
            postings = r.json()
            board_size = len(postings) if isinstance(postings, list) else None
            for j in postings:
                title = j.get("text", "")
                if not _title_maybe_relevant(title):
                    continue
                dt = None
                created_ms = j.get("createdAt", 0)
                if created_ms and created_ms > 0:
                    try:
                        dt = datetime.fromtimestamp(created_ms / 1000.0, tz=timezone.utc)
                    except (ValueError, OSError):
                        pass
                if dt is None:
                    dt = datetime.now(timezone.utc) - timedelta(hours=72)
                if hours_ago(dt) > max_h:
                    continue
                loc = j.get("categories", {}).get("location", "")
                out.append(job(title, name, loc, j.get("hostedUrl", ""), dt, "Lever",
                               j.get("descriptionPlain", ""), board_size=board_size))
        except Exception as e:
            print(f"WARNING: [Lever:{slug}] {e}")
    return out


def scrape_ashby(max_h):
    """Ashby's public REST job-board API. Startup-heavy, with real timestamps."""
    out = []
    for slug in ASHBY_SLUGS:
        try:
            r = requests.get(
                f"https://api.ashbyhq.com/posting-api/job-board/{slug}",
                headers=HEADERS, timeout=8,
            )
            if r.status_code != 200:
                continue
            postings = r.json().get("jobs", []) or []
            board_size = len(postings)
            for j in postings:
                if j.get("isListed") is False:
                    continue
                title = j.get("title", "")
                if not _title_maybe_relevant(title):
                    continue
                dt = None
                ts = j.get("publishedAt") or j.get("updatedAt") or ""
                if ts:
                    try:
                        dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
                    except ValueError:
                        pass
                if dt is None:
                    # Same fallback as Greenhouse/Lever: an undated posting is
                    # treated as 72h old, not silently dropped.
                    dt = datetime.now(timezone.utc) - timedelta(hours=72)
                if hours_ago(dt) > max_h:
                    continue
                loc = j.get("location") or ("Remote" if j.get("isRemote") else "")
                url = j.get("jobUrl") or j.get("applyUrl") or ""
                out.append(job(title, slug.replace("-", " ").title(), loc, url, dt,
                               "Ashby", j.get("descriptionPlain", ""), board_size=board_size))
        except Exception as e:
            print(f"WARNING: [Ashby:{slug}] {e}")
    return out


def scrape_remoteok(max_h):
    """RemoteOK exposes a documented JSON API at /api. Single polite request."""
    out = []
    try:
        r = requests.get("https://remoteok.com/api", headers=HEADERS, timeout=12)
        if r.status_code != 200:
            return out
        for j in r.json():
            if not isinstance(j, dict) or not j.get("position") or not j.get("company"):
                continue
            title = j.get("position", "")
            if not _title_maybe_relevant(title):
                continue
            dt = None
            epoch = j.get("epoch")
            if epoch:
                try:
                    dt = datetime.fromtimestamp(int(epoch), tz=timezone.utc)
                except (ValueError, OSError):
                    pass
            if dt is None and j.get("date"):
                try:
                    dt = datetime.fromisoformat(str(j["date"]).replace("Z", "+00:00"))
                except ValueError:
                    pass
            if dt is None:
                # Same fallback as Greenhouse/Lever: an undated posting is
                # treated as 72h old, not silently dropped.
                dt = datetime.now(timezone.utc) - timedelta(hours=72)
            if hours_ago(dt) > max_h:
                continue
            tags = j.get("tags", []) or []
            loc = j.get("location") or "Remote"
            desc = BeautifulSoup(j.get("description", ""), "html.parser").get_text()
            desc = (desc + " " + " ".join(tags)).strip()
            url = j.get("url") or j.get("apply_url", "")
            out.append(job(title, j.get("company", ""), loc, url, dt, "RemoteOK", desc))
    except Exception as e:
        print(f"WARNING: [RemoteOK] {e}")
    return out


def scrape_wellfound(max_h=None):
    """Best-effort HTML scrape. Off by default; fragile and unstable markup."""
    out = []
    try:
        r = requests.get("https://wellfound.com/jobs?role=software-engineer&remote=true",
                         headers=HEADERS, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        for card in soup.select("div[class*='JobListing'], div[class*='job-listing']")[:30]:
            a = card.find("a", href=True)
            co = card.find(class_=re.compile("company", re.I))
            if a and _title_maybe_relevant(a.text):
                dt = datetime.now(timezone.utc) - timedelta(hours=48)
                out.append(job(a.text.strip(), co.text.strip() if co else "",
                               "Remote", "https://wellfound.com" + a["href"], dt, "Wellfound"))
    except Exception as e:
        print(f"WARNING: [Wellfound] {e}")
    return out


def scrape_remotive(max_h):
    """Remotive's documented public API (remote-only board). Carries an
    employment-type field (contract/freelance/part_time) and a US-eligibility
    field, so it is the main remote/contract supply. Single polite request."""
    out = []
    try:
        r = requests.get("https://remotive.com/api/remote-jobs?category=software-dev",
                         headers=HEADERS, timeout=12)
        if r.status_code != 200:
            print(f"WARNING: [Remotive] HTTP {r.status_code}")
            return out
        for j in r.json().get("jobs", []) or []:
            title = j.get("title", "")
            if not _title_maybe_relevant(title):
                continue
            region = (j.get("candidate_required_location") or "").strip()
            rl = region.lower()
            if rl and not any(k in rl for k in ["usa", "united states", "americas",
                                                "worldwide", "anywhere", "north america"]):
                continue  # remote, but not US-eligible
            dt = None
            pub = j.get("publication_date")
            if pub:
                try:
                    dt = datetime.fromisoformat(str(pub).replace("Z", "+00:00"))
                except ValueError:
                    pass
            if dt is None:
                # Same fallback as Greenhouse/Lever: an undated posting is
                # treated as 72h old, not silently dropped.
                dt = datetime.now(timezone.utc) - timedelta(hours=72)
            if hours_ago(dt) > max_h:
                continue
            desc = BeautifulSoup(j.get("description", ""), "html.parser").get_text()
            loc = f"Remote ({region})" if region else "Remote"
            jd = job(title, j.get("company_name", ""), loc, j.get("url", ""),
                     dt, "Remotive", desc)
            # The API's employment type is authoritative; it beats title inference.
            hint = {"contract": "contract", "freelance": "contract",
                    "part_time": "part-time", "internship": "intern"}.get(j.get("job_type", ""))
            if hint:
                jd["role_type_hint"] = hint
            out.append(jd)
    except Exception as e:
        print(f"WARNING: [Remotive] {e}")
    return out


def scrape_yc(max_h):
    """YC Work at a Startup. robots.txt permits all crawling; listings are
    embedded server-side as Inertia data-page JSON (one polite request; the
    server requires a browser-style Accept header or it returns 406).
    Postings carry no timestamp, so a conservative 48h age is assumed
    (Wellfound precedent) -- YC roles never flag Fresh and are skipped
    entirely by windows under 48h such as --fresh."""
    out = []
    yc_headers = dict(HEADERS)
    yc_headers["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    dt = datetime.now(timezone.utc) - timedelta(hours=48)  # no per-job timestamp
    if hours_ago(dt) > max_h:
        return out
    # Two queries: the default eng set skews SF, remote=yes fills the remote
    # supply. Deduped by job id; polite pause between requests.
    urls = [
        "https://www.workatastartup.com/jobs?role=eng",
        "https://www.workatastartup.com/jobs?role=eng&remote=yes",
    ]
    seen = set()
    for i, u in enumerate(urls):
        if i:
            time.sleep(0.5)
        try:
            r = requests.get(u, headers=yc_headers, timeout=12)
            if r.status_code != 200:
                print(f"WARNING: [YC WaaS] HTTP {r.status_code}")
                continue
            m = re.search(r'data-page="([^"]+)"', r.text)
            if not m:
                print("WARNING: [YC WaaS] no embedded job data found (markup changed?)")
                continue
            payload = json.loads(html.unescape(m.group(1)))
            for j in payload.get("props", {}).get("jobs", []) or []:
                jid = j.get("id")
                if jid in seen:
                    continue
                seen.add(jid)
                title = j.get("title", "")
                role_typ = j.get("roleType", "")
                if role_typ in ("Hardware", "Mechanical", "Engineering manager"):
                    continue
                if j.get("jobType") == "Intern" and not config.INCLUDE_INTERNSHIPS:
                    continue
                if not _title_maybe_relevant(title):
                    continue
                url = f"https://www.workatastartup.com/jobs/{jid}" if jid else ""
                desc = " ".join(x for x in (role_typ, j.get("companyOneLiner", ""),
                                            j.get("salary") or "",
                                            "internship" if j.get("jobType") == "Intern" else "") if x)
                out.append(job(title, j.get("companyName", ""), j.get("location", ""),
                               url, dt, "YC WaaS", desc))
        except Exception as e:
            print(f"WARNING: [YC WaaS] {e}")
    return out


def scrape_jobspy(max_h):
    """Disabled by default (config.SOURCES). LinkedIn/Indeed via JobSpy."""
    out = []
    try:
        from jobspy import scrape_jobs
    except ImportError:
        print("  [jobspy] not installed; skipping")
        return out
    for term in ["AI engineer remote", "machine learning engineer", "full stack engineer python"]:
        try:
            df = scrape_jobs(site_name=["linkedin", "indeed"], search_term=term,
                             location="United States", results_wanted=15,
                             hours_old=max_h, country_indeed="USA", is_remote=True)
            if df is None or df.empty:
                continue
            for _, row in df.iterrows():
                title = str(row.get("title", ""))
                if not _title_maybe_relevant(title):
                    continue
                dt = row.get("date_posted")
                if hasattr(dt, "to_pydatetime"):
                    dt = dt.to_pydatetime()
                if isinstance(dt, datetime) and dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                if not isinstance(dt, datetime):
                    dt = datetime.now(timezone.utc) - timedelta(hours=72)
                if hours_ago(dt) > max_h:
                    continue
                out.append(job(title, str(row.get("company", "")), str(row.get("location", "")),
                               str(row.get("job_url", "")), dt, str(row.get("site", "")).title(),
                               str(row.get("description", ""))))
        except Exception as e:
            print(f"WARNING: [JobSpy:{term!r}] {e}")
    return out


# ===========================================================================
# Core pipeline: collect -> process (filter/rank/dedup)
# ===========================================================================

def collect_jobs(hours):
    """Scrape every enabled source. Returns (jobs, per-source counts)."""
    registry = [
        ("Greenhouse ATS",  "greenhouse", scrape_greenhouse, {"max_h": hours}),
        ("Lever ATS",       "lever",      scrape_lever,      {"max_h": hours}),
        ("Ashby ATS",       "ashby",      scrape_ashby,      {"max_h": hours}),
        ("RemoteOK",        "remoteok",   scrape_remoteok,   {"max_h": hours}),
        ("Remotive",        "remotive",   scrape_remotive,   {"max_h": hours}),
        ("YC Work at a Startup", "yc",    scrape_yc,         {"max_h": hours}),
        ("Wellfound",       "wellfound",  scrape_wellfound,  {"max_h": hours}),
        ("LinkedIn/Indeed", "jobspy",     scrape_jobspy,     {"max_h": hours}),
    ]
    jobs, stats = [], {}
    for label, key, fn, kw in registry:
        if not config.SOURCES.get(key, False):
            continue
        print(f"Scraping {label}...")
        found = fn(**kw)
        print(f"  {len(found)} candidate roles")
        jobs.extend(found)
        stats[label] = len(found)
    return jobs, stats


def process_jobs(raw, extra_keywords=None, target_locations=None):
    """Filter to relevant paid US/remote roles, rank, and dedupe across sources.

    extra_keywords/target_locations are optional per-run overrides from the web
    UI; passed explicitly (never via module globals) so concurrent runs stay
    isolated.
    """
    extra_keywords = [k.strip().lower() for k in (extra_keywords or []) if k and k.strip()]
    processed = []
    for j in raw:
        title, desc = j.get("title", ""), j.get("description", "")
        if not (is_relevant(title, desc, extra_keywords) or title_allowed(title)):
            continue
        if not title_on_track(title, extra_keywords):
            continue
        if is_excluded_company(j.get("company", "")):
            continue
        if is_excluded_function(title):
            continue
        if is_excluded_ops(title):
            continue
        if is_excluded_seniority(title):
            continue
        if years_required(desc) >= config.MAX_YEARS:
            continue
        if not is_usa_location(j.get("location", "")):
            continue
        if not matches_target_locations(j.get("location", ""), target_locations):
            continue
        mode = location_accept(j.get("location", ""))
        if mode is None:
            continue
        j["work_mode"] = mode
        # A remote-eligible role must read as remote everywhere (UI + Excel),
        # never as an onsite role in a city we don't accept ("San Francisco,
        # CA (Hybrid) OR Remote (...)" led with San Francisco).
        if mode == "remote" and not j.get("location", "").lower().startswith("remote"):
            j["location"] = f"Remote / {j['location']}"
        rt = j.get("role_type_hint") or role_type(title, desc)
        j["notes"] = ""
        if rt == "intern":
            if not config.INCLUDE_INTERNSHIPS:
                continue
            keep, rt, note = internship_guard(title, desc)
            if not keep:
                continue
            j["notes"] = note
        j["role_type"] = rt
        j["requires_degree"] = requires_degree(title, desc)
        if j["requires_degree"] and config.DEGREE_REQUIREMENT_MODE == "filter":
            continue
        j["rank_score"] = rank_score(j, extra_keywords)
        j["score"] = int(j["rank_score"])  # kept for the web frontend
        j["fresh"] = fresh_flag(j.get("hours_ago", 999))
        processed.append(j)

    processed.sort(key=lambda x: (-x["rank_score"], x.get("hours_ago", 999)))
    seen, deduped = set(), []
    for j in processed:
        key = (j["company"].lower().strip(), j["title"].lower().strip())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(j)
    return deduped


# ===========================================================================
# Outreach (contact research uses web search per policy)
# ===========================================================================

def _extract_text(resp):
    """Join all text blocks from an Anthropic response (tool use returns many)."""
    parts = []
    for block in getattr(resp, "content", []) or []:
        if getattr(block, "type", None) == "text":
            parts.append(block.text)
    return "\n".join(parts).strip()


def find_contacts(company, role, client):
    """Research up to 3 real outreach targets. Always uses web search."""
    prompt = (
        f'Research and return up to 3 real people at "{company}" who are the best '
        f'outreach targets for a "{role}" role. Prioritize engineering hiring '
        f'managers and technical recruiters, then engineers on the relevant team. '
        f'Use web search to find current, real people and their public LinkedIn profiles. '
        f'Return ONLY a JSON array, no prose:\n'
        f'[{{"name": "", "title": "", "reason": "", '
        f'"linkedin_search_url": "https://www.linkedin.com/search/results/people/?keywords=...", '
        f'"guessed_email": null}}]'
    )
    try:
        r = client.messages.create(
            model=OUTREACH_MODEL, max_tokens=1200,
            # Sonnet 5 runs adaptive thinking by default; disable it so the
            # small max_tokens budget goes to the answer, not reasoning.
            thinking={"type": "disabled"},
            tools=[WEB_SEARCH_TOOL],
            messages=[{"role": "user", "content": prompt}],
        )
        text = re.sub(r"```json|```", "", _extract_text(r)).strip()
        m = re.search(r"\[.*\]", text, re.S)
        return json.loads(m.group(0)) if m else []
    except Exception as e:
        print(f"WARNING: [find_contacts:{company}] {e}")
        return []


def draft_message(company, role, url, contact, client, desc=""):
    """Draft a short, tailored outreach note from the profile in .env."""
    name = contact.get("name") or "there"
    snippet = (desc or "").strip()[:400]
    try:
        r = client.messages.create(
            model=OUTREACH_MODEL, max_tokens=600,
            thinking={"type": "disabled"},
            messages=[{"role": "user", "content": (
                f"Draft a short outreach note from {YOUR_NAME or 'the candidate'} to {name}, "
                f"{contact.get('title', '')} at {company}, about the {role} role ({url}).\n"
                + (f"Job description excerpt: {snippet}\n" if snippet else "")
                + f"Background: {MY_BACKGROUND}\n"
                f"Positioning: entry-level candidate, under 1 year of professional experience. "
                f"Be honest about that - lead with genuine interest and concrete projects, and "
                f"never imply seniority. Mention one specific thing about this role or company.\n"
                f"Keep it under 90 words, specific, warm, professional. No em-dashes, no emojis.\n"
                f'Return JSON only: {{"linkedin_message": "(under 280 chars)", "email_subject": "", "email_body": ""}}'
            )}],
        )
        text = re.sub(r"```json|```", "", _extract_text(r)).strip()
        m = re.search(r"\{.*\}", text, re.S)
        return json.loads(m.group(0)) if m else {}
    except Exception as e:
        print(f"WARNING: [draft_message:{company}] {e}")
        return {}


def _fallback_contact(company, role, researched=False):
    """A usable contact when live research is unavailable or came up empty.
    Says so explicitly instead of leaving a blank that looks broken."""
    if researched:
        name = "(no named contact found - use LinkedIn search)"
    else:
        name = "(research off this run - use LinkedIn search)"
    return {
        "name": name,
        "title": "Engineering Recruiter or Hiring Manager",
        "reason": f"Owns or influences hiring for {role} at {company}. Search, then connect.",
        "linkedin_search_url": "https://www.linkedin.com/search/results/people/?keywords="
                               + quote_plus(f"{company} recruiter engineering"),
        "guessed_email": None,
    }


# Role-family hooks for template drafts. Checked in order; first match wins,
# so AI/ML outranks the generic software family. Stack names here mirror the
# default MY_BACKGROUND blurb; personal facts stay in .env.
_FAMILY_HOOKS = [
    (("machine learning", " ml ", " ai ", "ai/", "llm", "genai", "research", "applied scientist"),
     "I have been building applied-AI projects hands-on (RAG pipelines, LLM integrations) and want to do that work on a real team"),
    (("backend", "back-end", "back end"),
     "backend work in Python and FastAPI is where I am strongest"),
    (("frontend", "front-end", "front end", " web "),
     "I enjoy building clean React and TypeScript interfaces on top of real APIs"),
    (("full stack", "full-stack", "fullstack", "founding engineer", "product engineer",
      "software", "developer", "member of technical staff"),
     "I like owning features end to end, frontend through API and database"),
]


def _role_family_hook(role):
    t = f" {(role or '').lower()} "
    for keys, hook in _FAMILY_HOOKS:
        if any(k in t for k in keys):
            return hook
    return "I build software end to end and pick things up fast"


def _template_draft(company, role):
    """A profile-based draft used when the Claude draft is unavailable.
    Varies by role family; positioning stays honestly entry-level."""
    name = YOUR_NAME or "the candidate"
    hook = _role_family_hook(role)
    bg = MY_BACKGROUND or ("I build full-stack and applied-AI features end to end "
                           "(Python, FastAPI, React, TypeScript).")
    li = (f"Hi, I'm {name}, an early-career engineer. I saw the {role} opening at {company} and "
          f"{hook}. Open to a quick chat about it?")
    body = (f"Hi,\n\nI came across the {role} role at {company} and wanted to reach out directly. "
            f"I am an early-career engineer, and {hook}. {bg} "
            f"I would love to learn more about the team and where I could contribute.\n\n"
            f"Best,\n{name}")
    return {"linkedin_message": li[:280], "email_subject": f"{role} at {company}", "email_body": body}


def build_outreach(jobs, client, top_n):
    """
    For the top-ranked roles, populate a contact and a tailored draft. Always
    produces a row: if live research or drafting is unavailable (no funded key),
    it falls back to a target-profile contact and a profile-based template.
    """
    outreach = []
    targets = jobs[:top_n]
    print(f"Preparing outreach for top {len(targets)} roles...")
    for j in targets:
        co, role, url = j.get("company", ""), j.get("title", ""), j.get("job_url", "")
        if not co or not role:
            continue
        print(f"  {co} - {role}")
        contacts = find_contacts(co, role, client) if client is not None else []
        contact = contacts[0] if contacts else _fallback_contact(co, role, researched=client is not None)
        msgs = (draft_message(co, role, url, contact, client, j.get("description", ""))
                if (client is not None and contacts) else {})
        if not msgs:
            msgs = _template_draft(co, role)
        outreach.append({
            "company": co, "role": role, "rank_score": j.get("rank_score", 0),
            "name": contact.get("name", ""), "title": contact.get("title", ""),
            "reason": contact.get("reason", ""),
            "linkedin_url": contact.get("linkedin_search_url", ""),
            "linkedin_msg": msgs.get("linkedin_message", ""),
            "email": contact.get("guessed_email", "") or "",
            "email_subj": msgs.get("email_subject", ""),
            "email_body": msgs.get("email_body", ""),
        })
    return outreach


# ===========================================================================
# Excel output
# ===========================================================================

def write_excel(jobs, outreach):
    """Write the color-coded tracker: Jobs, Outreach, and Legend sheets.
    Returns the workbook filename."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    fname = f"job_leads_{ts}.xlsx"
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0F2940")

    # --- Jobs sheet ---
    ws = wb.active
    ws.title = "Jobs"
    headers = ["#", "Role", "Company", "Type", "Location", "Source", "Posted",
               "Fresh", "Rank", "Apply Link", "Applied?", "Status", "Notes"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for i, j in enumerate(jobs, 1):
        ha = j.get("hours_ago", 999)
        posted = f"{ha}h ago" if ha < 999 else "Unknown"
        ws.append([
            i, j["title"], j["company"], j.get("role_type", ""), j["location"],
            j["source"], posted, j.get("fresh", ""), j.get("rank_score", 0),
            j["job_url"], "", "Not Applied", j.get("notes", ""),
        ])
        color = (
            "F1948A" if ha < 1 else      # red: under 1 hour, act now
            "D5F5E3" if ha < 6 else      # green: under 6 hours
            "D6EAF8" if ha < 24 else     # blue: under 24 hours
            "FEF9E7" if ha < 48 else     # yellow: under 48 hours
            "FDFEFE"                      # white: older
        )
        for c in ws[i + 1]:
            c.fill = PatternFill("solid", fgColor=color)

    for i, w in enumerate([4, 40, 20, 15, 20, 12, 11, 7, 7, 50, 10, 14, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # --- Outreach sheet (ranked by leverage) ---
    ws2 = wb.create_sheet("Outreach")
    headers2 = ["Company", "Role", "Rank", "Contact Name", "Contact Title",
                "Why Reach Out", "LinkedIn Search", "Draft Message", "Guessed Email"]
    ws2.append(headers2)
    for c in ws2[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for row in outreach:
        draft = row.get("linkedin_msg", "") or row.get("email_body", "")
        ws2.append([
            row.get("company", ""), row.get("role", ""), row.get("rank_score", 0),
            row.get("name", ""), row.get("title", ""), row.get("reason", ""),
            row.get("linkedin_url", ""), draft, row.get("email", ""),
        ])
    for i, w in enumerate([20, 34, 7, 20, 24, 34, 45, 55, 28], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # --- Legend ---
    ws3 = wb.create_sheet("Legend")
    for row in [
        ["COLOR KEY", ""],
        ["Red", "Posted under 1 hour. Apply now, you may be first."],
        ["Green", "Posted under 6 hours. Apply today."],
        ["Blue", "Posted under 24 hours."],
        ["Yellow", "Posted under 48 hours."],
        ["White", "Older than 48 hours."],
        ["", ""],
        ["RANK", "Higher is a better match: profile keywords + freshness + paid role type."],
        ["TYPE", "full-time, contract-to-hire, or contract. Internships filtered out by default."],
        ["", ""],
        ["RUN MODES", ""],
        ["Standard", "python job_pipeline_full.py"],
        ["Fresh poll", "python job_pipeline_full.py --fresh"],
        ["Custom window", "python job_pipeline_full.py --hours 24"],
        ["Scrape only", "python job_pipeline_full.py --scrape-only"],
    ]:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 80

    # Caller saves the workbook so the cold-outreach module can append first.
    return fname, wb


# ===========================================================================
# API entry point (used by api.py)
# ===========================================================================

def run_pipeline_api(api_key, hours=None, cold_outreach_enabled=True, cold_outreach_limit=10,
                     target_locations=None, role_keywords=None, skills=None):
    """Run the pipeline and return structured JSON. Key used only for this call."""
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    hours = hours or config.DEFAULT_HOURS

    print(f"\n{'='*60}")
    print(f"  Pipeline run - {datetime.now().strftime('%Y-%m-%d %H:%M')} ({hours}h window)")
    print(f"  Contact research: ENABLED (key supplied per request)")
    print(f"{'='*60}\n")

    extra_keywords = (role_keywords or []) + (skills or [])

    raw, _ = collect_jobs(hours)
    jobs = process_jobs(raw, extra_keywords=extra_keywords, target_locations=target_locations)
    fresh = sum(1 for j in jobs if j.get("hours_ago", 999) <= 24)
    print(f"\nTotal: {len(jobs)} ranked roles | {fresh} posted in last 24h\n")

    outreach = build_outreach(jobs, client, config.OUTREACH_TOP_N) if jobs else []

    # Cold outreach runs standalone via cold_outreach.py; the in-pipeline
    # integration was removed. cold_outreach_enabled/cold_outreach_limit are
    # still accepted for API compatibility, and the empty list keeps the web
    # UI's Cold Outreach tab rendering its empty state.
    cold_contacts = []

    return {
        "jobs": jobs,
        "outreach": outreach,
        "cold_outreach": cold_contacts,
        "summary": {
            "total_jobs": len(jobs),
            "fresh_jobs": fresh,
            "outreach_count": len(outreach),
            "cold_outreach_count": len(cold_contacts),
        },
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# ===========================================================================
# CLI
# ===========================================================================

def main():
    """CLI entry point: parse flags, run the pipeline, write the Excel tracker."""
    p = argparse.ArgumentParser(description="EarlyBird job search pipeline")
    p.add_argument("--hours", type=int, default=config.DEFAULT_HOURS,
                   help="Lookback window in hours")
    p.add_argument("--fresh", action="store_true",
                   help="Fast poll: short window (config.FRESH_HOURS), scrape + Excel only")
    p.add_argument("--scrape-only", action="store_true",
                   help="Scrape and write Excel; skip all Claude API calls")
    p.add_argument("--limit", type=int, default=None, help="Cap number of jobs")
    p.add_argument("--no-email", action="store_true",
                   help="Deprecated no-op; outreach is draft-only and never sends mail")
    args = p.parse_args()

    fresh_mode = args.fresh
    hours = config.FRESH_HOURS if fresh_mode else args.hours
    do_claude = not args.scrape_only and not fresh_mode

    client = None
    if do_claude:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            print("Set ANTHROPIC_API_KEY in .env (or run with --scrape-only / --fresh)")
            sys.exit(1)
        for var in ["YOUR_NAME", "MY_BACKGROUND"]:
            if not os.getenv(var):
                print(f"WARNING: {var} not set in .env; outreach drafts may be generic")
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

    mode = "FRESH poll" if fresh_mode else ("scrape-only" if args.scrape_only else "standard")
    if client is not None:
        research_line = "ENABLED (live web research + tailored drafts)"
    elif fresh_mode:
        research_line = "DISABLED (--fresh mode: outreach skipped entirely)"
    else:
        research_line = "DISABLED (scrape-only: fallback contacts + template drafts)"
    print(f"\n{'='*60}")
    print(f"  EarlyBird - {datetime.now().strftime('%Y-%m-%d %H:%M')} ({mode})")
    print(f"  Companies: {len(COMPANIES)} | Window: {hours}h")
    print(f"  Contact research: {research_line}")
    print(f"{'='*60}\n")

    raw, stats = collect_jobs(hours)
    jobs = process_jobs(raw)
    if args.limit:
        jobs = jobs[:args.limit]

    print("\nPer-source (candidates scraped):")
    for label, n in stats.items():
        print(f"  {label}: {n}")
    fresh = sum(1 for j in jobs if j.get("hours_ago", 999) <= 24)
    urgent = sum(1 for j in jobs if j.get("hours_ago", 999) < 6)
    print(f"\nTotal: {len(jobs)} ranked roles | {fresh} in last 24h | {urgent} under 6h\n")

    # Outreach populates in standard mode (tailored via Claude, with fallback)
    # and in scrape-only mode (template drafts, no API key needed). --fresh
    # skips it for speed.
    outreach = []
    if not fresh_mode and jobs:
        outreach = build_outreach(jobs, client, config.OUTREACH_TOP_N)

    fname, wb = write_excel(jobs, outreach)
    wb.save(fname)

    if not fresh_mode:
        try:
            refresh_watchlist_cache()
        except Exception as e:
            print(f"WARNING: [watchlist refresh] {e}")

    print("\n" + "=" * 60)
    print("  COMPLETE")
    print(f"  {len(jobs)} roles | {fresh} fresh | {len(outreach)} outreach drafted")
    print(f"  File: {fname}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
