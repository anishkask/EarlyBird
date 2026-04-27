#!/usr/bin/env python3
"""
Early-Bird Job Hunter
=====================
Finds SWE internship opportunities BEFORE they hit LinkedIn/Indeed by querying
ATS platforms (Greenhouse, Lever, Ashby) directly, scraping VC job boards, and
searching YC Work at a Startup.

Outputs a multi-sheet Excel workbook:
  Sheet 1 - Jobs           : All discovered positions + application links
  Sheet 2 - Outreach       : Hiring manager contacts (LinkedIn + email)
  Sheet 3 - Cold Outreach  : Seed-stage companies with no open positions
  Sheet 4 - Dashboard      : Summary stats + quick links

Usage:
  python job_hunter.py

Optional API keys (free tiers available — see CONFIG below):
  Apollo.io  → contact/email enrichment
  Hunter.io  → email finding by domain
  SerpAPI    → Google Jobs search fallback
"""

import os
import sys
import time
import json
import re
import csv
import io
from datetime import datetime, timedelta
from urllib.parse import urlencode, quote_plus, urlparse

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    from openpyxl.comments import Comment
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl requests beautifulsoup4 --quiet")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — edit these before running
# ─────────────────────────────────────────────────────────────────────────────
CONFIG = {
    # ── Role keywords to match against job titles ──────────────────────────
    "keywords": [
        "software engineering intern",
        "software developer intern",
        "swe intern",
        "backend engineer intern",
        "frontend engineer intern",
        "fullstack intern",
        "full stack intern",
        "ai engineering intern",
        "machine learning intern",
        "data science intern",
        "data engineering intern",
        "platform engineer intern",
        "web developer intern",
        "product engineer intern",
        "computer science intern",
        "engineering intern",
        "developer intern",
        "data analyst intern",
        "it intern",
        "cloud intern",
        "security engineer intern",
        "automation intern",
    ],

    # ── Internship year (used to filter results) ───────────────────────────
    "year": "2026",

    # ── Optional API keys (leave empty string "" to skip that source) ──────
    "apollo_api_key": "",     # https://app.apollo.io  → free: 50 credits/mo
    "hunter_api_key": "",     # https://hunter.io      → free: 25 searches/mo
    "serpapi_key":    "",     # https://serpapi.com    → free: 100 searches/mo

    # ── Output ─────────────────────────────────────────────────────────────
    "output_file": "early_bird_jobs.xlsx",

    # ── Behaviour ──────────────────────────────────────────────────────────
    "max_per_ats_company": 30,   # Max jobs fetched per company slug
    "request_delay": 0.4,        # Seconds between requests (be polite)
    "timeout": 10,               # HTTP timeout in seconds
}

# ─────────────────────────────────────────────────────────────────────────────
# COMPANY LISTS  (ATS platform slugs)
# Add / remove company slugs freely — these are the URL identifiers used by
# each ATS.  Format: lowercase, hyphen-separated, no spaces.
# ─────────────────────────────────────────────────────────────────────────────

# ── Companies using Greenhouse ─────────────────────────────────────────────
GREENHOUSE_COMPANIES = [
    # AI / Foundation Models
    "anthropic", "openai", "scale", "cohere", "adept",
    "characterai", "runway", "elevenlabs", "perplexityai",
    "together-ai", "mistral-ai", "huggingface",
    # Dev Tools / Infra
    "vercel", "netlify", "supabase", "neon", "render",
    "doppler", "tailscale", "sentry", "datadog", "axiom",
    "linear", "shortcut", "height-app",
    # Fintech
    "stripe", "brex", "ramp", "mercury-technologies",
    "modern-treasury", "lithic", "robinhood", "coinbase", "kraken",
    # B2B SaaS
    "notion", "airtable", "retool", "webflow",
    "figma", "miro", "loom", "coda",
    "rippling", "gusto", "lattice", "culture-amp",
    "hubspot", "intercom", "mixpanel", "amplitude", "heap",
    "segment", "hightouch", "census",
    # Consumer / Marketplace
    "airbnb", "instacart", "doordash", "reddit", "discord",
    "roblox", "unity",
    # Climate / Mobility
    "waymo", "cruise", "aurora", "samsara", "motive",
    # Health
    "ro", "springhealth", "nomi-health", "color",
    # Defence / Space
    "anduril", "shieldai", "palantir",
]

# ── Companies using Lever ──────────────────────────────────────────────────
LEVER_COMPANIES = [
    "netflix", "dropbox", "pinterest", "quora", "zendesk",
    "plaid", "affirm", "klarna", "marqeta",
    "deel", "remote", "ripple", "polygon", "consensys",
    "gitlab", "hashicorp", "confluent", "cockroachdb",
    "clickhouse", "elastic", "algolia", "weaviate",
    "weights-biases", "lightning-ai",
    "joby", "archer", "rivian",
]

# ── Companies using Ashby ──────────────────────────────────────────────────
ASHBY_COMPANIES = [
    "linear", "mercury", "ramp", "moderntreasury",
    "column", "pilot-com", "bench-accounting",
    "clerk", "stytch",
    "plasmic", "builderdotio",
    "resend", "cal",
    "zed", "cursor", "codeium",
    "warp",
]

# ── Seed-stage / early-stage companies for COLD OUTREACH ─────────────────
# (No open intern posting yet — reach out to founder/CEO directly)
COLD_OUTREACH_COMPANIES = [
    # Recent YC batches (W24/S24/W25)
    {"name": "Letta",          "domain": "letta.ai",         "stage": "YC W24",   "focus": "AI agents"},
    {"name": "Cognition AI",   "domain": "cognition.ai",     "stage": "Seed",     "focus": "AI coding agent"},
    {"name": "E2B",            "domain": "e2b.dev",          "stage": "YC W23",   "focus": "AI code execution"},
    {"name": "Morph",          "domain": "morph.so",         "stage": "Seed",     "focus": "AI code editing"},
    {"name": "Lovable",        "domain": "lovable.dev",      "stage": "Seed",     "focus": "AI app builder"},
    {"name": "Bolt.new",       "domain": "bolt.new",         "stage": "Seed",     "focus": "AI full-stack builder"},
    {"name": "Windsurf",       "domain": "codeium.com",      "stage": "Series B", "focus": "AI IDE"},
    {"name": "Magic Dev",      "domain": "magic.dev",        "stage": "Seed",     "focus": "LLM for code"},
    {"name": "Factory AI",     "domain": "factory.ai",       "stage": "Seed",     "focus": "AI dev automation"},
    {"name": "Cosine",         "domain": "cosine.sh",        "stage": "Seed",     "focus": "AI coding assistant"},
    {"name": "Augment Code",   "domain": "augmentcode.com",  "stage": "Series A", "focus": "AI coding"},
    {"name": "Continue.dev",   "domain": "continue.dev",     "stage": "Seed",     "focus": "Open-source AI IDE"},
    {"name": "Cline",          "domain": "cline.bot",        "stage": "Seed",     "focus": "AI coding agent"},
    {"name": "Eko AI",         "domain": "eko.ai",           "stage": "Seed",     "focus": "AI automation"},
    {"name": "Induced AI",     "domain": "induced.ai",       "stage": "YC W23",   "focus": "AI browser automation"},
    {"name": "Browserbase",    "domain": "browserbase.com",  "stage": "Seed",     "focus": "Browser infrastructure"},
    {"name": "Stagehand",      "domain": "stagehand.dev",    "stage": "YC W25",   "focus": "AI browser agent"},
    {"name": "Firecrawl",      "domain": "firecrawl.dev",    "stage": "YC W24",   "focus": "AI web scraping"},
    {"name": "Exa AI",         "domain": "exa.ai",           "stage": "Series A", "focus": "AI search API"},
    {"name": "Tavily",         "domain": "tavily.com",       "stage": "Seed",     "focus": "AI search API"},
    {"name": "Composio",       "domain": "composio.dev",     "stage": "YC W24",   "focus": "AI tool integrations"},
    {"name": "LlamaIndex",     "domain": "llamaindex.ai",    "stage": "Series A", "focus": "LLM data framework"},
    {"name": "Langchain",      "domain": "langchain.com",    "stage": "Series A", "focus": "LLM framework"},
    {"name": "Helicone",       "domain": "helicone.ai",      "stage": "YC W23",   "focus": "LLM observability"},
    {"name": "Braintrust",     "domain": "braintrustdata.com","stage": "Seed",    "focus": "LLM evals"},
    {"name": "Confident AI",   "domain": "confident-ai.com", "stage": "Seed",     "focus": "LLM evals"},
    {"name": "Vapi",           "domain": "vapi.ai",          "stage": "YC W23",   "focus": "Voice AI API"},
    {"name": "Bland AI",       "domain": "bland.ai",         "stage": "Seed",     "focus": "AI phone calls"},
    {"name": "Retell AI",      "domain": "retellai.com",     "stage": "YC W24",   "focus": "Voice AI"},
    {"name": "Hume AI",        "domain": "hume.ai",          "stage": "Series B", "focus": "Emotional AI"},
]

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html, */*",
})


def safe_get(url, params=None, headers=None, retries=2):
    for attempt in range(retries + 1):
        try:
            r = SESSION.get(
                url, params=params, headers=headers,
                timeout=CONFIG["timeout"]
            )
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            if attempt < retries:
                time.sleep(1.5)
            else:
                return None
    return None


def keyword_match(title: str) -> bool:
    t = title.lower()
    # Must contain "intern" or "co-op" or "coop"
    if not any(k in t for k in ["intern", "co-op", "coop", "co op"]):
        return False
    # Match at least one of our keywords
    keywords = CONFIG["keywords"]
    year = CONFIG["year"]
    # Year filter — if the title includes a year, it must match ours
    year_pattern = re.search(r"20\d\d", t)
    if year_pattern and year_pattern.group() != year:
        return False
    for kw in keywords:
        parts = kw.lower().split()
        if all(p in t for p in parts):
            return True
    # Broad fallback: any engineering/developer/tech intern title
    tech_words = ["engineer", "developer", "software", "data", "ml", "ai",
                  "backend", "frontend", "fullstack", "cloud", "platform",
                  "devops", "security", "automation", "analytics", "product"]
    return any(w in t for w in tech_words)


def clean_text(text):
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def domain_from_url(url):
    try:
        return urlparse(url).netloc.replace("www.", "")
    except Exception:
        return ""


# ─────────────────────────────────────────────────────────────────────────────
# JOB FETCHING — ATS PLATFORMS
# ─────────────────────────────────────────────────────────────────────────────

def fetch_greenhouse(slug):
    """Query Greenhouse public jobs API — no auth required."""
    url = f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs"
    r = safe_get(url, params={"content": "true"})
    jobs = []
    if not r:
        return jobs
    try:
        data = r.json()
    except Exception:
        return jobs
    for j in data.get("jobs", []):
        title = j.get("title", "")
        if not keyword_match(title):
            continue
        location = ""
        loc_list = j.get("offices") or j.get("location", {})
        if isinstance(loc_list, list) and loc_list:
            location = loc_list[0].get("name", "")
        elif isinstance(loc_list, dict):
            location = loc_list.get("name", "")
        jobs.append({
            "company":      slug.replace("-", " ").title(),
            "title":        title,
            "location":     location,
            "posted_date":  _parse_gh_date(j.get("updated_at", "")),
            "apply_url":    j.get("absolute_url", ""),
            "source":       "Greenhouse (Direct ATS)",
            "ats":          "Greenhouse",
            "company_url":  f"https://www.greenhouse.io/company/{slug}",
            "domain":       f"{slug.replace('-', '')}.com",
        })
        time.sleep(0.05)
    return jobs


def _parse_gh_date(s):
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%b %d, %Y")
    except Exception:
        return ""


def fetch_lever(slug):
    """Query Lever public postings API — no auth required."""
    url = f"https://api.lever.co/v0/postings/{slug}"
    r = safe_get(url, params={"mode": "json", "limit": CONFIG["max_per_ats_company"]})
    jobs = []
    if not r:
        return jobs
    try:
        postings = r.json()
    except Exception:
        return jobs
    if not isinstance(postings, list):
        return jobs
    for j in postings:
        title = j.get("text", "")
        if not keyword_match(title):
            continue
        loc_data = j.get("categories", {})
        location = loc_data.get("location", "") or loc_data.get("allLocations", [""])[0]
        ts = j.get("createdAt", 0)
        posted = ""
        if ts:
            try:
                posted = datetime.fromtimestamp(ts / 1000).strftime("%b %d, %Y")
            except Exception:
                pass
        jobs.append({
            "company":      slug.replace("-", " ").title(),
            "title":        title,
            "location":     location,
            "posted_date":  posted,
            "apply_url":    j.get("hostedUrl", ""),
            "source":       "Lever (Direct ATS)",
            "ats":          "Lever",
            "company_url":  f"https://jobs.lever.co/{slug}",
            "domain":       f"{slug.replace('-', '')}.com",
        })
    return jobs


def fetch_ashby(slug):
    """Query Ashby public job board API — no auth required."""
    url = "https://api.ashbyhq.com/posting-public/job-board/list"
    r = safe_get(url, params={"organizationHostedJobsPageName": slug})
    jobs = []
    if not r:
        return jobs
    try:
        data = r.json()
    except Exception:
        return jobs
    for j in data.get("jobPostings", []):
        title = j.get("title", "")
        if not keyword_match(title):
            continue
        location = j.get("locationName", "") or j.get("isRemote", "")
        if j.get("isRemote"):
            location = location or "Remote"
        jobs.append({
            "company":      j.get("organizationName", slug.replace("-", " ").title()),
            "title":        title,
            "location":     location,
            "posted_date":  _parse_gh_date(j.get("publishedDate", "")),
            "apply_url":    j.get("jobUrl", f"https://jobs.ashbyhq.com/{slug}/{j.get('id','')}"),
            "source":       "Ashby (Direct ATS)",
            "ats":          "Ashby",
            "company_url":  f"https://jobs.ashbyhq.com/{slug}",
            "domain":       f"{slug.replace('-', '')}.com",
        })
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
# JOB FETCHING — YC WORK AT A STARTUP
# ─────────────────────────────────────────────────────────────────────────────

def fetch_yc_jobs():
    """Scrape YC Work at a Startup internship listings."""
    print("  → Fetching YC Work at a Startup...")
    url = "https://www.workatastartup.com/jobs"
    params = {
        "jobType": "internship",
        "role":    "eng",
    }
    r = safe_get(url, params=params)
    jobs = []
    if not r:
        return jobs

    soup = BeautifulSoup(r.text, "html.parser")
    # Each job card
    for card in soup.select("div.job-name, div[data-job-id]")[:80]:
        title_el = card.select_one("a.job-name, span.job-name, h2")
        if not title_el:
            continue
        title = clean_text(title_el.get_text())
        if not keyword_match(title):
            continue
        company_el = card.select_one("a.company-name, span.company-name")
        company = clean_text(company_el.get_text()) if company_el else "YC Company"
        link_el = card.select_one("a[href]")
        href = link_el["href"] if link_el else ""
        apply_url = href if href.startswith("http") else f"https://www.workatastartup.com{href}"
        jobs.append({
            "company":      company,
            "title":        title,
            "location":     "Varies",
            "posted_date":  "",
            "apply_url":    apply_url,
            "source":       "YC Work at a Startup",
            "ats":          "YC",
            "company_url":  apply_url,
            "domain":       "",
        })
    # Fallback: try the JSON API endpoint (available on some YC pages)
    if not jobs:
        api_url = "https://www.workatastartup.com/company_positions/list_all"
        r2 = safe_get(api_url, params={"jobType": "internship"})
        if r2:
            try:
                data = r2.json()
                for item in data.get("results", []):
                    title = item.get("title", "")
                    if not keyword_match(title):
                        continue
                    jobs.append({
                        "company":      item.get("company_name", ""),
                        "title":        title,
                        "location":     item.get("remote", "Varies"),
                        "posted_date":  "",
                        "apply_url":    item.get("url", ""),
                        "source":       "YC Work at a Startup",
                        "ats":          "YC",
                        "company_url":  item.get("company_url", ""),
                        "domain":       domain_from_url(item.get("company_url", "")),
                    })
            except Exception:
                pass
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
# JOB FETCHING — VC PORTFOLIO JOB BOARDS
# ─────────────────────────────────────────────────────────────────────────────

def fetch_a16z_jobs():
    print("  → Fetching a16z portfolio jobs...")
    url = "https://jobs.a16z.com/jobs"
    r = safe_get(url, params={"query": "software intern", "count": 50})
    return _parse_lever_talent_board(r, "a16z Portfolio")


def fetch_first_round_jobs():
    print("  → Fetching First Round Capital jobs...")
    url = "https://jobs.firstround.com/jobs"
    r = safe_get(url, params={"query": "intern", "count": 50})
    return _parse_lever_talent_board(r, "First Round Capital Portfolio")


def _parse_lever_talent_board(r, source_name):
    """Many VC job boards are powered by Lever's Talent Hub — parse their JSON."""
    jobs = []
    if not r:
        return jobs
    try:
        data = r.json()
        postings = data.get("postings", data) if isinstance(data, dict) else data
        if not isinstance(postings, list):
            return jobs
        for j in postings:
            title = j.get("text", j.get("title", ""))
            if not keyword_match(title):
                continue
            jobs.append({
                "company":      j.get("company", {}).get("name", source_name),
                "title":        title,
                "location":     j.get("categories", {}).get("location", ""),
                "posted_date":  "",
                "apply_url":    j.get("hostedUrl", j.get("applyUrl", "")),
                "source":       source_name,
                "ats":          "Lever Talent Hub",
                "company_url":  j.get("company", {}).get("url", ""),
                "domain":       "",
            })
    except Exception:
        pass
    return jobs


def fetch_simplify_jobs():
    """Simplify.jobs aggregates internship listings — search their open dataset."""
    print("  → Fetching Simplify.jobs internship listings...")
    # Simplify exposes a public CSV/JSON of SWE internships on GitHub
    url = "https://raw.githubusercontent.com/SimplifyJobs/Summer2026-Internships/dev/.github/scripts/listings.json"
    r = safe_get(url)
    jobs = []
    if not r:
        return jobs
    try:
        data = r.json()
        for j in data:
            title = j.get("title", "")
            if not keyword_match(title):
                continue
            # Each listing may have multiple links
            links = j.get("url", "")
            if isinstance(links, list):
                links = links[0] if links else ""
            jobs.append({
                "company":      j.get("company_name", ""),
                "title":        title,
                "location":     ", ".join(j.get("locations", [])),
                "posted_date":  j.get("date_posted", ""),
                "apply_url":    links,
                "source":       "Simplify.jobs (GitHub)",
                "ats":          "Various",
                "company_url":  f"https://simplify.jobs/c/{j.get('company_slug','')}",
                "domain":       domain_from_url(links),
            })
    except Exception:
        pass
    return jobs


def fetch_pitt_csc_jobs():
    """PittCSC Summer Internship list — widely maintained GitHub repo."""
    print("  → Fetching PittCSC internship list (GitHub)...")
    url = "https://raw.githubusercontent.com/pittcsc/Summer2026-Internships/dev/README.md"
    r = safe_get(url)
    jobs = []
    if not r:
        return jobs
    # Parse Markdown table rows
    for line in r.text.split("\n"):
        if "|" not in line or line.strip().startswith("|---") or "Company" in line:
            continue
        cols = [c.strip() for c in line.split("|") if c.strip()]
        if len(cols) < 3:
            continue
        company_raw = cols[0]
        title_raw   = cols[1]
        location    = cols[2] if len(cols) > 2 else ""
        # Extract hyperlinks [text](url) from markdown
        company_match = re.search(r"\[(.+?)\]\((.+?)\)", company_raw)
        company     = company_match.group(1) if company_match else clean_text(re.sub(r"\[.*?\]|\(.*?\)", "", company_raw))
        title_match = re.search(r"\[(.+?)\]\((.+?)\)", title_raw)
        title       = title_match.group(1) if title_match else clean_text(title_raw)
        apply_url   = title_match.group(2) if title_match else (company_match.group(2) if company_match else "")
        if not company or not title:
            continue
        if not keyword_match(title):
            continue
        jobs.append({
            "company":      company,
            "title":        title,
            "location":     clean_text(location),
            "posted_date":  "",
            "apply_url":    apply_url,
            "source":       "PittCSC GitHub List",
            "ats":          "Various",
            "company_url":  "",
            "domain":       domain_from_url(apply_url),
        })
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
# OPTIONAL: Google Jobs via SerpAPI
# ─────────────────────────────────────────────────────────────────────────────

def fetch_serpapi_jobs():
    key = CONFIG.get("serpapi_key", "")
    if not key:
        return []
    print("  → Searching Google Jobs via SerpAPI...")
    jobs = []
    queries = [
        "software engineering intern 2026 startup",
        "SWE intern 2026 site:greenhouse.io OR site:lever.co OR site:ashbyhq.com",
        "AI engineering intern 2026",
        "data science intern 2026 YC startup",
    ]
    for q in queries:
        params = {
            "engine": "google_jobs",
            "q":      q,
            "api_key": key,
            "hl":     "en",
            "num":    20,
        }
        r = safe_get("https://serpapi.com/search", params=params)
        if not r:
            continue
        try:
            data = r.json()
        except Exception:
            continue
        for j in data.get("jobs_results", []):
            title = j.get("title", "")
            if not keyword_match(title):
                continue
            links = j.get("related_links", [])
            apply_url = links[0].get("link", "") if links else j.get("share_link", "")
            jobs.append({
                "company":      j.get("company_name", ""),
                "title":        title,
                "location":     j.get("location", ""),
                "posted_date":  j.get("detected_extensions", {}).get("posted_at", ""),
                "apply_url":    apply_url,
                "source":       "Google Jobs (SerpAPI)",
                "ats":          "Various",
                "company_url":  "",
                "domain":       "",
            })
        time.sleep(CONFIG["request_delay"])
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
# CONTACT RESEARCH
# ─────────────────────────────────────────────────────────────────────────────

def build_linkedin_search_url(company, role="engineering recruiter OR hiring manager"):
    q = f"{company} {role}"
    return f"https://www.linkedin.com/search/results/people/?keywords={quote_plus(q)}"


def guess_email_patterns(first, last, domain):
    """Return common email format guesses for a person + company domain."""
    f, l = first.lower(), last.lower()
    return [
        f"{f}.{l}@{domain}",
        f"{f}{l}@{domain}",
        f"{f[0]}{l}@{domain}",
        f"{f}@{domain}",
        f"{f[0]}.{l}@{domain}",
    ]


def find_contacts_apollo(company_name, domain):
    """
    Search Apollo.io for engineering recruiters / hiring managers.
    Requires APOLLO_API_KEY in CONFIG.
    Free tier: 50 credits/month.
    """
    key = CONFIG.get("apollo_api_key", "")
    if not key:
        return []
    url = "https://api.apollo.io/v1/people/search"
    payload = {
        "api_key":            key,
        "q_organization_name": company_name,
        "person_titles":       [
            "Engineering Recruiter", "Technical Recruiter",
            "Recruiting Manager", "Head of Engineering",
            "VP Engineering", "Engineering Manager",
            "CTO", "Co-Founder", "Founder",
        ],
        "page":  1,
        "per_page": 5,
    }
    try:
        r = requests.post(url, json=payload, timeout=CONFIG["timeout"])
        data = r.json()
        contacts = []
        for p in data.get("people", []):
            contacts.append({
                "name":     f"{p.get('first_name','')} {p.get('last_name','')}".strip(),
                "title":    p.get("title", ""),
                "email":    p.get("email", ""),
                "linkedin": p.get("linkedin_url", ""),
            })
        return contacts
    except Exception:
        return []


def find_email_hunter(domain):
    """
    Search Hunter.io for email addresses at a domain.
    Requires HUNTER_API_KEY in CONFIG.
    Free tier: 25 searches/month.
    """
    key = CONFIG.get("hunter_api_key", "")
    if not key:
        return []
    url = "https://api.hunter.io/v2/domain-search"
    params = {
        "domain":  domain,
        "api_key": key,
        "limit":   5,
        "type":    "personal",
    }
    try:
        r = safe_get(url, params=params)
        if not r:
            return []
        data = r.json()
        contacts = []
        for email_data in data.get("data", {}).get("emails", []):
            name = f"{email_data.get('first_name','')} {email_data.get('last_name','')}".strip()
            contacts.append({
                "name":     name,
                "title":    email_data.get("position", ""),
                "email":    email_data.get("value", ""),
                "linkedin": "",
            })
        return contacts
    except Exception:
        return []


def find_contacts(company_name, domain):
    """Try Apollo → Hunter → generate LinkedIn search URL as fallback."""
    contacts = find_contacts_apollo(company_name, domain)
    if not contacts and domain:
        contacts = find_email_hunter(domain)
    # Always add a LinkedIn search URL fallback
    li_url = build_linkedin_search_url(company_name)
    if not contacts:
        contacts = [{
            "name":     "Search on LinkedIn →",
            "title":    "Engineering Recruiter / Hiring Manager",
            "email":    "",
            "linkedin": li_url,
        }]
    else:
        for c in contacts:
            if not c.get("linkedin"):
                c["linkedin"] = build_linkedin_search_url(company_name, c.get("title", ""))
    return contacts


# ─────────────────────────────────────────────────────────────────────────────
# DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def deduplicate(jobs):
    seen = set()
    unique = []
    for j in jobs:
        key = (j["company"].lower().strip(), j["title"].lower().strip())
        if key not in seen:
            seen.add(key)
            unique.append(j)
    return unique


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATION
# ─────────────────────────────────────────────────────────────────────────────

# ── Colour palette ─────────────────────────────────────────────────────────
C_NAVY      = "1A1A2E"
C_PURPLE    = "6C63FF"
C_TEAL      = "00C9A7"
C_GOLD      = "FFD700"
C_LIGHT_BG  = "F4F6FF"
C_ALT_ROW   = "EFF1FF"
C_WHITE     = "FFFFFF"
C_DARK_TEXT = "1A1A2E"
C_MED_GRAY  = "B0B3C6"
C_GREEN     = "2ECC71"
C_RED       = "E74C3C"
C_ORANGE    = "F39C12"

FONT_MAIN = "Arial"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_DARK_TEXT, size=11, italic=False):
    return Font(name=FONT_MAIN, bold=bold, color=color, size=size, italic=italic)

def _border_thin():
    s = Side(border_style="thin", color="D0D3E8")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, headers, row=1,
                bg=C_NAVY, fg=C_WHITE, bold=True, size=11):
    for col, txt in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = _font(bold=bold, color=fg, size=size)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border_thin()

def _data_row(ws, row_num, values, alt=False):
    bg = C_ALT_ROW if alt else C_WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font      = _font()
        c.fill      = _fill(bg)
        c.alignment = _align(wrap=True)
        c.border    = _border_thin()

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _hyperlink_cell(ws, row, col, url, display, bg=C_WHITE):
    c = ws.cell(row=row, column=col, value=display)
    if url:
        c.hyperlink = url
        c.font = Font(name=FONT_MAIN, color=C_PURPLE, underline="single", size=11)
    else:
        c.font = _font()
    c.fill      = _fill(bg)
    c.alignment = _align(wrap=True)
    c.border    = _border_thin()


def _title_block(ws, title, subtitle, col_span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = _font(bold=True, color=C_WHITE, size=16)
    c.fill      = _fill(C_NAVY)
    c.alignment = _align("center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font      = _font(italic=True, color=C_PURPLE, size=11)
    c.fill      = _fill(C_LIGHT_BG)
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22


# ─── Sheet 1: Jobs ───────────────────────────────────────────────────────────

def build_jobs_sheet(wb, jobs):
    ws = wb.active
    ws.title = "🔍 Jobs"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    N_COLS = 9
    _title_block(ws,
        "🚀 Early-Bird Internship Tracker",
        f"Generated {datetime.now().strftime('%B %d, %Y')}  ·  {len(jobs)} positions found",
        N_COLS)

    headers = [
        "Company", "Role / Title", "Location", "Posted Date",
        "Apply Link", "Source", "ATS Platform", "Status", "Notes"
    ]
    _header_row(ws, headers, row=4, bg=C_PURPLE)

    for i, j in enumerate(jobs, 1):
        row = i + 4
        alt = i % 2 == 0
        bg  = C_ALT_ROW if alt else C_WHITE

        _data_row(ws, row, ["", "", "", "", "", "", "", "", ""], alt=alt)
        ws.cell(row=row, column=1, value=j["company"]).font = _font(bold=True)
        ws.cell(row=row, column=1).fill = _fill(bg)
        ws.cell(row=row, column=1).alignment = _align()
        ws.cell(row=row, column=1).border = _border_thin()

        ws.cell(row=row, column=2, value=j["title"]).fill = _fill(bg)
        ws.cell(row=row, column=2).font = _font()
        ws.cell(row=row, column=2).alignment = _align(wrap=True)
        ws.cell(row=row, column=2).border = _border_thin()

        ws.cell(row=row, column=3, value=j.get("location", "")).fill = _fill(bg)
        ws.cell(row=row, column=3).font = _font(color=C_DARK_TEXT)
        ws.cell(row=row, column=3).alignment = _align()
        ws.cell(row=row, column=3).border = _border_thin()

        ws.cell(row=row, column=4, value=j.get("posted_date", "")).fill = _fill(bg)
        ws.cell(row=row, column=4).font = _font()
        ws.cell(row=row, column=4).alignment = _align("center")
        ws.cell(row=row, column=4).border = _border_thin()

        _hyperlink_cell(ws, row, 5, j.get("apply_url", ""), "Apply →", bg=bg)

        for col_i, val in enumerate([j["source"], j["ats"], "Not Applied", ""], 1):
            c = ws.cell(row=row, column=5 + col_i, value=val)
            c.fill = _fill(bg)
            c.font = _font()
            c.alignment = _align("center" if col_i < 3 else "left")
            c.border = _border_thin()

        # Colour-code Status
        status_cell = ws.cell(row=row, column=8)
        status_cell.value = "Not Applied"

    # Status dropdown via data validation
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1='"Not Applied,Applied,Interview,Offer,Rejected"',
            allow_blank=True, showErrorMessage=False
        )
        ws.add_data_validation(dv)
        dv.sqref = f"H5:H{len(jobs)+4}"
    except Exception:
        pass

    _set_col_widths(ws, [26, 38, 22, 16, 14, 28, 18, 16, 28])
    ws.row_dimensions[4].height = 22
    for i in range(len(jobs)):
        ws.row_dimensions[i + 5].height = 22


# ─── Sheet 2: Outreach ────────────────────────────────────────────────────────

def build_outreach_sheet(wb, jobs_with_contacts):
    ws = wb.create_sheet("📬 Outreach")
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    N_COLS = 10
    _title_block(ws,
        "📬 Outreach Tracker — Companies With Open Positions",
        "Find the right person. Send the right message. Beat everyone else.",
        N_COLS)

    headers = [
        "Company", "Role Applied For", "Contact Name", "Title",
        "LinkedIn URL", "Email", "Email Confidence", "Outreach Status",
        "Last Contact Date", "Notes / Draft"
    ]
    _header_row(ws, headers, row=4, bg=C_TEAL, fg=C_NAVY)

    row = 5
    for j in jobs_with_contacts:
        contacts = j.get("contacts", [])
        if not contacts:
            contacts = [{"name": "", "title": "", "email": "", "linkedin": ""}]
        alt = (row % 2 == 0)
        bg  = C_ALT_ROW if alt else C_WHITE

        for ci, contact in enumerate(contacts[:2]):  # max 2 contacts per job
            _data_row(ws, row, [""] * N_COLS, alt=(row % 2 == 0))
            bg = C_ALT_ROW if (row % 2 == 0) else C_WHITE

            # Company (merge across contacts visually with same row colouring)
            c1 = ws.cell(row=row, column=1, value=j["company"] if ci == 0 else "")
            c1.font = _font(bold=(ci == 0))
            c1.fill = _fill(bg); c1.alignment = _align(); c1.border = _border_thin()

            c2 = ws.cell(row=row, column=2, value=j["title"] if ci == 0 else "")
            c2.font = _font(); c2.fill = _fill(bg); c2.alignment = _align(wrap=True)
            c2.border = _border_thin()

            name_cell = ws.cell(row=row, column=3, value=contact.get("name",""))
            name_cell.font = _font(bold=True)
            name_cell.fill = _fill(bg); name_cell.alignment = _align()
            name_cell.border = _border_thin()

            title_cell = ws.cell(row=row, column=4, value=contact.get("title",""))
            title_cell.font = _font(); title_cell.fill = _fill(bg)
            title_cell.alignment = _align(); title_cell.border = _border_thin()

            li = contact.get("linkedin", "")
            _hyperlink_cell(ws, row, 5, li, "View Profile →" if li else "", bg=bg)

            email = contact.get("email", "")
            _hyperlink_cell(ws, row, 6,
                f"mailto:{email}" if email else "",
                email or "—", bg=bg)

            # Email confidence indicator
            conf = "✅ Verified" if email and "@" in email else "🔍 Search needed"
            c7 = ws.cell(row=row, column=7, value=conf)
            c7.font = _font(color=C_GREEN if "✅" in conf else C_ORANGE)
            c7.fill = _fill(bg); c7.alignment = _align("center")
            c7.border = _border_thin()

            for col_i, val in enumerate(["Not Started", "", ""], 1):
                cx = ws.cell(row=row, column=7 + col_i, value=val)
                cx.font = _font(); cx.fill = _fill(bg)
                cx.alignment = _align("center" if col_i < 3 else "left", wrap=True)
                cx.border = _border_thin()

            row += 1

    # Dropdown
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1='"Not Started,Message Sent,Replied,Scheduled Call,No Response"',
            allow_blank=True, showErrorMessage=False
        )
        ws.add_data_validation(dv)
        dv.sqref = f"H5:H{row}"
    except Exception:
        pass

    _set_col_widths(ws, [24, 34, 24, 26, 20, 32, 18, 20, 18, 40])
    ws.row_dimensions[4].height = 22


# ─── Sheet 3: Cold Outreach ──────────────────────────────────────────────────

def build_cold_outreach_sheet(wb, cold_companies):
    ws = wb.create_sheet("🧊 Cold Outreach")
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    N_COLS = 11
    _title_block(ws,
        "🧊 Cold Outreach — Seed-Stage Companies (No Open Intern Role)",
        "Reach out before a role exists. Be the first face they think of.",
        N_COLS)

    headers = [
        "Company", "Domain", "Stage / Funding", "Focus Area",
        "Founder / CEO Name", "Founder LinkedIn", "Founder Email",
        "LinkedIn Search", "Email Guess #1", "Outreach Status", "Notes"
    ]
    _header_row(ws, headers, row=4, bg=C_NAVY, fg=C_GOLD)

    for i, co in enumerate(cold_companies, 1):
        row = i + 4
        alt = i % 2 == 0
        bg  = C_ALT_ROW if alt else C_WHITE

        _data_row(ws, row, [""] * N_COLS, alt=alt)

        domain = co.get("domain", "")
        li_search = build_linkedin_search_url(co["name"], "founder OR CEO OR co-founder")
        email_guess = f"founder@{domain}" if domain else ""

        vals = [
            co["name"], domain, co.get("stage",""),
            co.get("focus",""), "", "", "",
        ]
        for col_i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_i, value=v)
            c.font = _font(bold=(col_i == 1))
            c.fill = _fill(bg); c.alignment = _align(wrap=True)
            c.border = _border_thin()

        _hyperlink_cell(ws, row, 8, li_search, "Search LinkedIn →", bg=bg)

        eg = ws.cell(row=row, column=9, value=email_guess)
        eg.font = _font(italic=True, color=C_MED_GRAY)
        eg.fill = _fill(bg); eg.alignment = _align(); eg.border = _border_thin()

        for col_i, v in enumerate(["Not Started", ""], 1):
            cx = ws.cell(row=row, column=9 + col_i, value=v)
            cx.font = _font(); cx.fill = _fill(bg)
            cx.alignment = _align("center" if col_i == 1 else "left", wrap=True)
            cx.border = _border_thin()

    # Dropdown
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1='"Not Started,Message Sent,Replied,Scheduled Call,No Response"',
            allow_blank=True, showErrorMessage=False
        )
        ws.add_data_validation(dv)
        dv.sqref = f"J5:J{len(cold_companies)+4}"
    except Exception:
        pass

    _set_col_widths(ws, [22, 24, 18, 28, 24, 20, 28, 20, 28, 18, 36])
    ws.row_dimensions[4].height = 22


# ─── Sheet 4: Dashboard ─────────────────────────────────────────────────────

def build_dashboard_sheet(wb, jobs, cold_companies):
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines = False

    # Background
    for r in range(1, 50):
        for c_i in range(1, 12):
            cell = ws.cell(row=r, column=c_i)
            cell.fill = _fill(C_LIGHT_BG)

    # Title
    ws.merge_cells("B2:J3")
    t = ws["B2"]
    t.value = "🚀  Early-Bird Job Hunt Dashboard"
    t.font  = Font(name=FONT_MAIN, bold=True, color=C_NAVY, size=22)
    t.alignment = _align("center")

    ws.merge_cells("B4:J4")
    s = ws["B4"]
    s.value = f"Last updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}  ·  Focus: SWE Internships 2026"
    s.font  = _font(italic=True, color=C_PURPLE, size=11)
    s.alignment = _align("center")

    ws.row_dimensions[2].height = 40
    ws.row_dimensions[4].height = 20

    # ── KPI Tiles ────────────────────────────────────────────────────────────
    kpis = [
        ("Total Positions", len(jobs),         C_PURPLE, "B"),
        ("Sources Scraped",  5,                C_TEAL,   "D"),
        ("Cold Outreach",   len(cold_companies),C_NAVY,  "F"),
        ("Applied",         0,                 C_GREEN,  "H"),
    ]
    for label, val, color, col in kpis:
        ws.merge_cells(f"{col}6:{col}7")
        num_cell = ws[f"{col}6"]
        num_cell.value     = val
        num_cell.font      = Font(name=FONT_MAIN, bold=True, color=C_WHITE, size=28)
        num_cell.fill      = _fill(color)
        num_cell.alignment = _align("center")
        ws.row_dimensions[6].height = 40
        ws.row_dimensions[7].height = 40

        ws.merge_cells(f"{col}8:{col}8")
        lbl_cell = ws[f"{col}8"]
        lbl_cell.value     = label
        lbl_cell.font      = _font(bold=True, color=color, size=11)
        lbl_cell.alignment = _align("center")
        ws.row_dimensions[8].height = 20

    # ── Quick Navigation ─────────────────────────────────────────────────────
    ws.merge_cells("B10:J10")
    nav_title = ws["B10"]
    nav_title.value     = "Quick Navigation"
    nav_title.font      = _font(bold=True, color=C_NAVY, size=13)
    nav_title.alignment = _align("left")
    ws.row_dimensions[10].height = 24

    nav_links = [
        ("B11", "🔍 View All Jobs",         "'🔍 Jobs'!A1"),
        ("D11", "📬 Outreach Tracker",      "'📬 Outreach'!A1"),
        ("F11", "🧊 Cold Outreach",         "'🧊 Cold Outreach'!A1"),
    ]
    ws.row_dimensions[11].height = 28
    for cell_ref, label, link in nav_links:
        c = ws[cell_ref]
        c.value     = label
        c.hyperlink = f"#{link}"
        c.font      = Font(name=FONT_MAIN, bold=True, color=C_WHITE, size=11,
                           underline="single")
        c.fill      = _fill(C_PURPLE)
        c.alignment = _align("center")

    # ── Top sources breakdown ────────────────────────────────────────────────
    ws.merge_cells("B13:J13")
    src_title = ws["B13"]
    src_title.value = "Positions by Source"
    src_title.font  = _font(bold=True, color=C_NAVY, size=13)
    src_title.alignment = _align("left")
    ws.row_dimensions[13].height = 24

    from collections import Counter
    source_counts = Counter(j["source"] for j in jobs)
    row_i = 14
    for src, cnt in source_counts.most_common(10):
        ws.merge_cells(f"B{row_i}:G{row_i}")
        lbl = ws[f"B{row_i}"]
        lbl.value     = src
        lbl.font      = _font(size=10)
        lbl.alignment = _align("left")

        num = ws[f"H{row_i}"]
        num.value     = cnt
        num.font      = _font(bold=True, color=C_PURPLE, size=10)
        num.alignment = _align("center")
        num.fill      = _fill(C_ALT_ROW)

        ws.row_dimensions[row_i].height = 18
        row_i += 1

    # ── Tips box ─────────────────────────────────────────────────────────────
    tip_row = row_i + 2
    ws.merge_cells(f"B{tip_row}:J{tip_row}")
    t2 = ws.cell(row=tip_row, column=2, value="💡  Pro Tips")
    t2.font = _font(bold=True, color=C_NAVY, size=13)
    ws.row_dimensions[tip_row].height = 24

    tips = [
        "⚡ Greenhouse / Lever / Ashby postings are live 1–3 weeks BEFORE LinkedIn aggregates them.",
        "🧊 Cold outreach to seed-stage founders often works better than applying — they'll create a role for you.",
        "📬 Personalize every outreach: mention a specific project, blog post, or product feature you love.",
        "🔄 Re-run this script weekly to catch fresh postings before the inbox floods.",
        "🔑 Add Apollo / Hunter API keys in CONFIG for automatic email discovery.",
    ]
    for tip in tips:
        tip_row += 1
        ws.merge_cells(f"B{tip_row}:J{tip_row}")
        tc = ws.cell(row=tip_row, column=2, value=tip)
        tc.font      = _font(size=10)
        tc.alignment = _align(wrap=True)
        ws.row_dimensions[tip_row].height = 20

    # Column widths
    for col_letter, w in zip("ABCDEFGHIJK", [4, 22, 4, 22, 4, 22, 4, 14, 4, 4, 4]):
        ws.column_dimensions[col_letter].width = w


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*60)
    print("  🚀  Early-Bird Job Hunter")
    print("  Targeting: SWE Internships 2026")
    print("="*60 + "\n")

    all_jobs = []

    # ── 1. ATS platforms (Greenhouse, Lever, Ashby) ────────────────────────
    print("📡 Scanning ATS platforms...")
    ats_total = len(GREENHOUSE_COMPANIES) + len(LEVER_COMPANIES) + len(ASHBY_COMPANIES)
    done = 0

    for slug in GREENHOUSE_COMPANIES:
        jobs = fetch_greenhouse(slug)
        all_jobs.extend(jobs)
        done += 1
        if jobs:
            print(f"  ✓ Greenhouse: {slug:30s} → {len(jobs)} match(es)")
        time.sleep(CONFIG["request_delay"])

    for slug in LEVER_COMPANIES:
        jobs = fetch_lever(slug)
        all_jobs.extend(jobs)
        done += 1
        if jobs:
            print(f"  ✓ Lever:      {slug:30s} → {len(jobs)} match(es)")
        time.sleep(CONFIG["request_delay"])

    for slug in ASHBY_COMPANIES:
        jobs = fetch_ashby(slug)
        all_jobs.extend(jobs)
        done += 1
        if jobs:
            print(f"  ✓ Ashby:      {slug:30s} → {len(jobs)} match(es)")
        time.sleep(CONFIG["request_delay"])

    # ── 2. Community lists ──────────────────────────────────────────────────
    print("\n📋 Scanning community-maintained lists...")
    all_jobs.extend(fetch_simplify_jobs())
    all_jobs.extend(fetch_pitt_csc_jobs())
    all_jobs.extend(fetch_yc_jobs())

    # ── 3. VC portfolio boards ──────────────────────────────────────────────
    print("\n🏦 Scanning VC portfolio job boards...")
    all_jobs.extend(fetch_a16z_jobs())
    all_jobs.extend(fetch_first_round_jobs())

    # ── 4. SerpAPI (optional) ───────────────────────────────────────────────
    if CONFIG.get("serpapi_key"):
        print("\n🔍 Running Google Jobs search...")
        all_jobs.extend(fetch_serpapi_jobs())

    # ── 5. Deduplicate ──────────────────────────────────────────────────────
    all_jobs = deduplicate(all_jobs)
    print(f"\n✅ {len(all_jobs)} unique positions found after deduplication")

    # Sort by posted date (most recent first), then company
    def sort_key(j):
        try:
            return datetime.strptime(j["posted_date"], "%b %d, %Y")
        except Exception:
            return datetime.min
    all_jobs.sort(key=sort_key, reverse=True)

    # ── 6. Contact research ─────────────────────────────────────────────────
    print("\n👤 Researching contacts...")
    use_api = CONFIG.get("apollo_api_key") or CONFIG.get("hunter_api_key")
    jobs_with_contacts = []
    seen_companies = set()
    for j in all_jobs:
        company = j["company"]
        if company not in seen_companies:
            seen_companies.add(company)
            if use_api:
                print(f"  → {company}")
                contacts = find_contacts(company, j.get("domain", ""))
                time.sleep(CONFIG["request_delay"])
            else:
                # No API key — generate LinkedIn search URL only
                contacts = [{
                    "name":     "Search on LinkedIn →",
                    "title":    "Engineering Recruiter / Hiring Manager",
                    "email":    "",
                    "linkedin": build_linkedin_search_url(company),
                }]
            j["contacts"] = contacts
        else:
            j["contacts"] = []
        jobs_with_contacts.append(j)

    if not use_api:
        print("  ℹ️  No API keys set — LinkedIn search URLs generated as fallback.")
        print("     Add apollo_api_key or hunter_api_key to CONFIG for auto email discovery.")

    # ── 7. Build Excel ──────────────────────────────────────────────────────
    print("\n📊 Building Excel workbook...")
    wb = Workbook()

    build_dashboard_sheet(wb, all_jobs, COLD_OUTREACH_COMPANIES)
    build_jobs_sheet(wb, all_jobs)
    build_outreach_sheet(wb, jobs_with_contacts)
    build_cold_outreach_sheet(wb, COLD_OUTREACH_COMPANIES)

    out_path = CONFIG["output_file"]
    wb.save(out_path)
    print(f"\n🎉 Saved → {out_path}")
    print(f"   Sheets: Dashboard · Jobs ({len(all_jobs)}) · Outreach · Cold Outreach ({len(COLD_OUTREACH_COMPANIES)})\n")

    return out_path


if __name__ == "__main__":
    main()
