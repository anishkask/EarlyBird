#!/usr/bin/env python3
"""
Cold Outreach module for EarlyBird pipeline.

Scrapes Philadelphia-area VC portfolio pages for newly listed small software
companies, enriches each with CEO/founder contact info via Apollo.io, resolves
LinkedIn profiles, generates a personalized cold email via Claude API, and
writes everything to a "Cold Outreach" tab in the existing Excel workbook.

Exposed API:
    run_cold_outreach(workbook, config, gmail_service=None) -> list[dict]
    build_config() -> dict
"""

import os
import re
import json
import time
import logging
from datetime import date
from pathlib import Path
from urllib.parse import quote_plus, urlparse

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Data directories
# ---------------------------------------------------------------------------
DATA_DIR = Path("data")
SEEN_COMPANIES_FILE = DATA_DIR / "seen_companies.json"
APOLLO_CACHE_FILE = DATA_DIR / "apollo_cache.json"
DRAFTS_DIR = DATA_DIR / "drafts"


def _ensure_data_dirs():
    DATA_DIR.mkdir(exist_ok=True)
    DRAFTS_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def build_config() -> dict:
    """Build cold-outreach config dict from environment variables."""
    vc_env = os.getenv("COLD_OUTREACH_VC_SOURCES", "")
    if vc_env:
        vc_sources = [s.strip() for s in vc_env.split(",") if s.strip()]
    else:
        vc_sources = [
            "https://www.robinhoodventures.com/portfolio/",
            "https://osageventurepartners.com/portfolio/",
            "https://www.dreamit.com/portfolio-securetech",
            "https://www.dreamit.com/portfolio-healthtech",
        ]
    return {
        "cold_outreach_enabled": os.getenv("COLD_OUTREACH_ENABLED", "true").lower() == "true",
        "cold_outreach_max_per_day": int(os.getenv("COLD_OUTREACH_MAX_PER_DAY", "10")),
        "create_gmail_drafts": os.getenv("CREATE_GMAIL_DRAFTS", "true").lower() == "true",
        "apollo_api_key": os.getenv("APOLLO_API_KEY", ""),
        "cold_outreach_min_employees": int(os.getenv("COLD_OUTREACH_MIN_EMPLOYEES", "1")),
        "cold_outreach_max_employees": int(os.getenv("COLD_OUTREACH_MAX_EMPLOYEES", "75")),
        "cold_outreach_vc_sources": vc_sources,
        "anthropic_api_key": os.getenv("ANTHROPIC_API_KEY", ""),
    }


# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------

def _load_json(path, default):
    try:
        p = Path(path)
        if p.exists():
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        logger.warning(f"Could not load {path}: {e}")
    return default


def _save_json(path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logger.warning(f"Could not save {path}: {e}")


# ---------------------------------------------------------------------------
# PART 1: VC Portfolio Scraper
# ---------------------------------------------------------------------------

_SCRAPER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

_SKIP_DOMAINS = {
    "linkedin.com", "twitter.com", "x.com", "facebook.com", "instagram.com",
    "crunchbase.com", "angel.co", "github.com", "youtube.com", "medium.com",
    "techcrunch.com", "pitchbook.com", "bloomberg.com", "forbes.com",
}


def _robots_allowed(page_url: str) -> bool:
    """Return True if the URL is scrapable per robots.txt (fail-open)."""
    try:
        from urllib.robotparser import RobotFileParser
        parsed = urlparse(page_url)
        robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
        rp = RobotFileParser()
        rp.set_url(robots_url)
        rp.read()
        return rp.can_fetch("*", page_url)
    except Exception:
        return True


def _extract_companies_from_page(url: str, source_label: str) -> list:
    """
    Fetch one VC portfolio page and extract company dicts.
    Returns list of {"name": str, "website": str, "vc_source": str}.
    """
    companies = []
    try:
        if not _robots_allowed(url):
            logger.warning(f"[VC:{source_label}] robots.txt disallows scraping, skipping")
            return companies

        r = requests.get(url, headers=_SCRAPER_HEADERS, timeout=15)
        if r.status_code in (403, 429):
            logger.warning(f"[VC:{source_label}] HTTP {r.status_code} — skipping this source today")
            return companies
        if r.status_code != 200:
            logger.warning(f"[VC:{source_label}] HTTP {r.status_code}")
            return companies

        soup = BeautifulSoup(r.text, "html.parser")
        vc_domain = urlparse(url).netloc.replace("www.", "")
        seen_names: set = set()

        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href or href.startswith("#") or href.startswith("mailto:"):
                continue
            if not href.startswith("http"):
                continue

            link_domain = urlparse(href).netloc.replace("www.", "")
            # Skip VC's own domain
            if vc_domain in link_domain or link_domain in vc_domain:
                continue
            # Skip social / aggregator links
            if any(sd in link_domain for sd in _SKIP_DOMAINS):
                continue

            # Derive company name: img alt > link text > domain stem
            name = ""
            img = a.find("img")
            if img and img.get("alt", "").strip():
                name = img["alt"].strip()
            if not name:
                text = a.get_text(separator=" ", strip=True)
                if 2 <= len(text) <= 80:
                    name = text
            if not name:
                stem = link_domain.split(".")[0]
                if stem:
                    name = stem.title()

            if not name or len(name) < 2:
                continue

            name_key = name.lower().strip()
            if name_key in seen_names:
                continue
            seen_names.add(name_key)

            companies.append({
                "name": name,
                "website": href.rstrip("/"),
                "vc_source": url,
            })

        logger.info(f"[VC:{source_label}] Extracted {len(companies)} candidate companies")

    except requests.exceptions.RequestException as e:
        logger.warning(f"[VC:{source_label}] Request error: {e}")
    except Exception as e:
        logger.warning(f"[VC:{source_label}] Unexpected error: {e}")

    return companies


def _url_resolves(website: str) -> bool:
    """Return True if the company website returns HTTP 200."""
    try:
        r = requests.head(website, headers=_SCRAPER_HEADERS, timeout=8, allow_redirects=True)
        if r.status_code == 200:
            return True
        # Some servers don't handle HEAD; fall back to GET
        if r.status_code in (405, 403):
            r2 = requests.get(website, headers=_SCRAPER_HEADERS, timeout=8, stream=True)
            return r2.status_code == 200
        return False
    except Exception:
        return False


# ---------------------------------------------------------------------------
# PART 2: Apollo.io Contact Enrichment
# ---------------------------------------------------------------------------

def _construct_email_guesses(first_name: str, last_name: str, domain: str) -> list:
    """Return list of probable email addresses in priority order."""
    if not first_name or not domain:
        return []
    fn = first_name.lower().strip()
    ln = last_name.lower().strip() if last_name else ""
    guesses = [f"{fn}@{domain}"]
    if ln:
        guesses += [
            f"{fn}.{ln}@{domain}",
            f"{fn[0]}.{ln}@{domain}",
            f"{fn}{ln}@{domain}",
        ]
    return guesses


def enrich_via_apollo(
    company_name: str,
    website_url: str,
    api_key: str,
    min_emp: int = 1,
    max_emp: int = 75,
) -> dict | None:
    """
    Call Apollo.io People Search API to find the CEO/founder of a company.
    Results are cached in data/apollo_cache.json.
    Returns a contact dict or None.
    """
    if not api_key:
        logger.warning(f"[Apollo] No API key — skipping enrichment for {company_name}")
        return None

    cache = _load_json(APOLLO_CACHE_FILE, {})
    cache_key = company_name.lower().strip()

    if cache_key in cache:
        cached = cache[cache_key]
        if cached is None:
            return None
        emp = cached.get("employees")
        if emp and (emp < min_emp or emp > max_emp):
            logger.info(f"[Apollo] Cache: {company_name} has {emp} employees — outside range, skipping")
            return None
        logger.info(f"[Apollo] Cache hit: {company_name}")
        return cached

    try:
        payload = {
            "api_key": api_key,
            "q_organization_name": company_name,
            "person_titles": ["CEO", "Co-Founder", "Founder", "CTO"],
            "per_page": 1,
        }
        r = requests.post(
            "https://api.apollo.io/v1/people/search",
            headers={"Content-Type": "application/json", "Cache-Control": "no-cache"},
            json=payload,
            timeout=15,
        )

        if r.status_code == 429:
            logger.warning("[Apollo] Rate limit reached — stopping enrichment for today")
            return None
        if r.status_code == 401:
            logger.warning("[Apollo] Unauthorized — check APOLLO_API_KEY")
            return None
        if r.status_code != 200:
            logger.warning(f"[Apollo] HTTP {r.status_code} for {company_name}")
            cache[cache_key] = None
            _save_json(APOLLO_CACHE_FILE, cache)
            return None

        people = r.json().get("people", [])
        if not people:
            logger.info(f"[Apollo] No people found for {company_name}")
            cache[cache_key] = None
            _save_json(APOLLO_CACHE_FILE, cache)
            return None

        person = people[0]
        org = person.get("organization") or {}
        emp_count = org.get("estimated_num_employees")

        if emp_count and (emp_count < min_emp or emp_count > max_emp):
            logger.info(f"[Apollo] {company_name} has {emp_count} employees — outside [{min_emp},{max_emp}], skipping")
            cache[cache_key] = None
            _save_json(APOLLO_CACHE_FILE, cache)
            return None

        # Derive domain for email guessing
        site = org.get("website_url") or website_url or ""
        domain = urlparse(site).netloc.replace("www.", "") if site else ""

        email = person.get("email") or None
        email_status = "VERIFIED" if email else "NOT FOUND"
        if not email and domain:
            guesses = _construct_email_guesses(
                person.get("first_name", ""), person.get("last_name", ""), domain
            )
            if guesses:
                email = guesses[0]
                email_status = "ESTIMATED"

        contact = {
            "company_name": org.get("name") or company_name,
            "website_url": org.get("website_url") or website_url,
            "first_name": person.get("first_name", ""),
            "last_name": person.get("last_name", ""),
            "title": person.get("title", ""),
            "email": email,
            "email_status": email_status,
            "linkedin_url": person.get("linkedin_url") or "",
            "employees": emp_count,
            "domain": domain,
        }

        cache[cache_key] = contact
        _save_json(APOLLO_CACHE_FILE, cache)
        return contact

    except requests.exceptions.RequestException as e:
        logger.warning(f"[Apollo] Request error for {company_name}: {e}")
        return None
    except Exception as e:
        logger.warning(f"[Apollo] Unexpected error for {company_name}: {e}")
        return None


# ---------------------------------------------------------------------------
# PART 3: LinkedIn URL Resolution
# ---------------------------------------------------------------------------

def resolve_linkedin(
    first_name: str,
    last_name: str,
    company_name: str,
    apollo_linkedin_url: str = "",
) -> tuple[str, str]:
    """
    Return (linkedin_url, url_type) where url_type is "DIRECT" or "SEARCH".
    Never scrapes LinkedIn — uses Apollo URL if available, otherwise
    constructs a LinkedIn people-search URL as fallback.
    """
    if apollo_linkedin_url:
        return apollo_linkedin_url, "DIRECT"
    query = quote_plus(f"{first_name} {last_name} {company_name}".strip())
    search_url = f"https://www.linkedin.com/search/results/people/?keywords={query}"
    return search_url, "SEARCH"


# ---------------------------------------------------------------------------
# PART 4: Deduplication helpers
# ---------------------------------------------------------------------------

def _load_seen_companies() -> set:
    data = _load_json(SEEN_COMPANIES_FILE, [])
    if isinstance(data, list):
        return {c.lower().strip() for c in data if isinstance(c, str)}
    return set()


def _save_seen_companies(seen_set: set):
    _save_json(SEEN_COMPANIES_FILE, sorted(seen_set))


def _existing_emails_in_sheet(ws) -> set:
    """Collect all email values already present in column E of the sheet."""
    emails: set = set()
    if ws is None:
        return emails
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 5 and row[4]:
            emails.add(str(row[4]).lower().strip())
    return emails


# ---------------------------------------------------------------------------
# PART 6: Cold Email Generation via Claude API
# ---------------------------------------------------------------------------

_SYSTEM_PROMPT = (
    "You are drafting a cold outreach email on behalf of Anishka Kakade, a senior "
    "Computer Science student at Temple University in Philadelphia graduating December 2026. "
    "She is seeking a software engineering internship for Summer 2026. "
    "Her stack: Python, FastAPI, React, TypeScript, PostgreSQL, Docker, ChromaDB, OpenAI API. "
    "Key projects: semantic code search pipeline (RAG/OpenAI/ChromaDB/FastAPI/React), "
    "ReMo full-stack app (TypeScript/React/FastAPI/PostgreSQL/Google OAuth with real users), "
    "automated job search pipeline (scrapes 10+ sources, scores listings, drafts emails via Claude API). "
    "At Bourns Inc internship: built RPA pipelines + REST API integrations with SAP workflows. "
    "She volunteers with her service sorority on Philly neighborhood cleanups. "
    "She is a daily SEPTA rider (Market-Frankford line). "
    "She grew up in an immigrant family without access to financial advisors. "
    "Writing rules: never use em-dashes, professional but human tone, 3 paragraphs max, "
    "no flattery stats, lead with something specific about the company."
)


def generate_cold_email(contact: dict, api_key: str) -> str:
    """
    Call Claude API to generate a personalized cold email.
    Returns full draft text (subject line first) or empty string on failure.
    """
    if not api_key:
        logger.warning("[Claude] No ANTHROPIC_API_KEY — skipping draft generation")
        return ""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        user_prompt = (
            f"Write a cold email to {contact.get('first_name', '')} {contact.get('last_name', '')}, "
            f"{contact.get('title', '')} at {contact.get('company_name', '')}. "
            f"Their website is {contact.get('website_url', '')}. "
            f"They were found via {contact.get('vc_source', '')}. "
            f"Company size: ~{contact.get('employees', 'unknown')} employees. "
            f"Subject line included as first line in format 'Subject: ...' "
            f"Then a blank line, then the email body. "
            f"Sign off: Anishka Kakade / anishkakade.vercel.app | anishka.s.kakade@gmail.com"
        )
        r = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=600,
            system=_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_prompt}],
        )
        return r.content[0].text.strip()
    except Exception as e:
        logger.warning(f"[Claude] Draft failed for {contact.get('company_name', '')}: {e}")
        return ""


def _save_draft_file(company_name: str, draft_text: str) -> str:
    """Persist full draft to data/drafts/<slug>.txt. Returns path string."""
    try:
        slug = re.sub(r"[^\w\s-]", "", company_name.lower()).strip()
        slug = re.sub(r"[\s-]+", "_", slug)
        path = DRAFTS_DIR / f"{slug}.txt"
        path.write_text(draft_text, encoding="utf-8")
        return str(path)
    except Exception as e:
        logger.warning(f"[Drafts] Could not save draft for {company_name}: {e}")
        return ""


def _parse_subject_and_body(draft_text: str) -> tuple[str, str]:
    """Split 'Subject: ...\n\n<body>' format into (subject, body)."""
    subject = ""
    body_lines = []
    past_blank = False
    for line in draft_text.split("\n"):
        if not subject and line.startswith("Subject:"):
            subject = line.replace("Subject:", "", 1).strip()
        elif subject:
            if not past_blank and line.strip() == "":
                past_blank = True
            elif past_blank or line.strip():
                body_lines.append(line)
                past_blank = True
    return subject, "\n".join(body_lines).strip()


# ---------------------------------------------------------------------------
# PART 7: Gmail Draft Creation
# ---------------------------------------------------------------------------

def create_gmail_draft(
    gmail_service,
    to_email: str,
    subject: str,
    body: str,
    from_email: str = "",
) -> tuple[str | None, str | None]:
    """
    Create a Gmail draft.  Requires gmail.compose scope.
    Returns (draft_id, compose_url) or (None, None) on failure.
    """
    if not gmail_service or not to_email:
        return None, None
    try:
        import base64
        from email.mime.text import MIMEText

        msg = MIMEText(body, "plain")
        msg["to"] = to_email
        msg["subject"] = subject
        if from_email:
            msg["from"] = from_email

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        draft = gmail_service.users().drafts().create(
            userId="me", body={"message": {"raw": raw}}
        ).execute()
        draft_id = draft.get("id", "")
        compose_url = (
            f"https://mail.google.com/mail/u/0/#drafts?compose={draft_id}"
            if draft_id else ""
        )
        return draft_id, compose_url
    except Exception as e:
        logger.warning(
            f"[Gmail Drafts] Could not create draft for {to_email}: {e}. "
            "Tip: delete token.json and re-run to re-authorize with gmail.compose scope."
        )
        return None, None


# ---------------------------------------------------------------------------
# PART 5: Cold Outreach Excel Tab
# ---------------------------------------------------------------------------

SHEET_NAME = "Cold Outreach"
SHEET_NAME_TEST = "Cold Outreach TEST"

_HEADER_COLOR = "1E3A5F"
_ALT_ROW_COLOR = "F8FAFC"
_ESTIMATED_EMAIL_COLOR = "FEF9C3"
_LINK_COLOR = "0563C1"

_COL_HEADERS = [
    "#", "Company Name", "Contact Name", "Title", "Email", "Email Status",
    "LinkedIn URL", "Company Website", "Employees", "VC Source", "Date Added",
    "Email Draft", "Gmail Draft Link", "Outreach Status", "Response Date", "Notes",
]
_COL_WIDTHS = [4, 28, 22, 22, 32, 14, 20, 32, 10, 30, 14, 45, 20, 16, 16, 20]

_STATUS_OPTIONS = ["Not Sent", "Sent", "Replied", "Meeting Booked", "Pass", "No Response"]


def write_cold_outreach_sheet(wb, contacts: list, sheet_name: str = SHEET_NAME):
    """
    Add or update the Cold Outreach sheet in *wb*.
    Creates header row if the sheet is new; otherwise appends rows.
    Does not save the workbook — caller is responsible.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor=_HEADER_COLOR)
    alt_fill = PatternFill("solid", fgColor=_ALT_ROW_COLOR)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    est_fill = PatternFill("solid", fgColor=_ESTIMATED_EMAIL_COLOR)

    # Get or create sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Write header if sheet is empty
    is_new = ws.max_row == 1 and ws.cell(1, 1).value is None
    if is_new:
        ws.append(_COL_HEADERS)
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"

    # Determine next row number counter (column A)
    last_num = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, int) and val > last_num:
            last_num = val

    # Dropdown validation for Outreach Status column (N = col 14)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(_STATUS_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)

    for idx, contact in enumerate(contacts):
        row_counter = last_num + idx + 1
        row_fill = white_fill if row_counter % 2 == 0 else alt_fill

        li_url = contact.get("linkedin_url", "")
        website = contact.get("website_url", "")
        draft_text = contact.get("draft_text", "")
        draft_preview = (draft_text[:297] + "...") if len(draft_text) > 300 else draft_text

        ws.append([
            row_counter,
            contact.get("company_name", ""),
            f"{contact.get('first_name', '')} {contact.get('last_name', '')}".strip(),
            contact.get("title", ""),
            contact.get("email", ""),
            contact.get("email_status", ""),
            li_url,
            website,
            contact.get("employees", ""),
            contact.get("vc_source", ""),
            contact.get("date_added", str(date.today())),
            draft_preview,
            contact.get("gmail_draft_url", ""),
            "Not Sent",
            "",
            "",
        ])
        actual_row = ws.max_row

        # Row fill and alignment
        for col_idx, cell in enumerate(ws[actual_row], 1):
            cell.fill = row_fill
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == 12),  # wrap Email Draft column
            )

        # Email cell: yellow highlight if estimated
        email_cell = ws.cell(actual_row, 5)
        if contact.get("email_status") == "ESTIMATED":
            email_cell.fill = est_fill

        # LinkedIn hyperlink
        if li_url:
            cell = ws.cell(actual_row, 7)
            cell.value = "View Profile"
            cell.hyperlink = li_url
            cell.font = Font(color=_LINK_COLOR, underline="single")

        # Company website hyperlink
        if website:
            cell = ws.cell(actual_row, 8)
            cell.value = website
            cell.hyperlink = website
            cell.font = Font(color=_LINK_COLOR, underline="single")

        # Gmail draft hyperlink
        gmail_url = contact.get("gmail_draft_url", "")
        if gmail_url:
            cell = ws.cell(actual_row, 13)
            cell.value = "Open Draft"
            cell.hyperlink = gmail_url
            cell.font = Font(color=_LINK_COLOR, underline="single")

        # Apply dropdown to status cell
        dv.add(ws.cell(actual_row, 14))

    # Column widths (max 50)
    for col_idx, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 50)

    return ws


# ---------------------------------------------------------------------------
# PART 8: Main Entry Point
# ---------------------------------------------------------------------------

def run_cold_outreach(
    workbook,
    config: dict,
    gmail_service=None,
    test_mode: bool = False,
    test_source_limit: int = 1,
    test_company_limit: int = 2,
) -> list:
    """
    Main entry point for the Cold Outreach module.

    Args:
        workbook:           openpyxl Workbook — will have the Cold Outreach sheet added.
        config:             Dict from build_config() or equivalent.
        gmail_service:      Authenticated Gmail API service (for draft creation); None = skip.
        test_mode:          If True, writes to "Cold Outreach TEST" sheet and does NOT
                            update seen_companies.json.
        test_source_limit:  (test_mode only) number of VC sources to scrape.
        test_company_limit: (test_mode only) max companies to process.

    Returns:
        List of new contact dicts added this run.
    """
    _ensure_data_dirs()

    if not config.get("cold_outreach_enabled", True) and not test_mode:
        logger.info("[Cold Outreach] Disabled in config — skipping")
        return []

    sheet_name = SHEET_NAME_TEST if test_mode else SHEET_NAME
    max_per_day = config.get("cold_outreach_max_per_day", 10)
    min_emp = config.get("cold_outreach_min_employees", 1)
    max_emp = config.get("cold_outreach_max_employees", 75)
    apollo_key = config.get("apollo_api_key") or os.getenv("APOLLO_API_KEY", "")
    anthropic_key = config.get("anthropic_api_key") or os.getenv("ANTHROPIC_API_KEY", "")
    create_drafts = config.get("create_gmail_drafts", True)
    vc_sources = list(config.get("cold_outreach_vc_sources", []))

    if test_mode:
        vc_sources = vc_sources[:test_source_limit]

    # --- Deduplication seed data ---
    seen_companies = _load_seen_companies()
    existing_sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
    existing_emails = _existing_emails_in_sheet(existing_sheet)

    # === PART 1: Scrape VC portfolio pages ===
    print(f"\n[Cold Outreach] Scraping {len(vc_sources)} VC portfolio page(s)...")
    new_companies: list = []

    for vc_url in vc_sources:
        source_label = urlparse(vc_url).netloc.replace("www.", "")
        companies = _extract_companies_from_page(vc_url, source_label)
        before = len(new_companies)
        for company in companies:
            if company["name"].lower().strip() not in seen_companies:
                new_companies.append(company)
        added = len(new_companies) - before
        print(f"  {source_label}: {added} new companies")
        time.sleep(2)

    # Validate company URLs (HTTP 200)
    print(f"[Cold Outreach] Validating {len(new_companies)} company URLs...")
    validated: list = []
    for company in new_companies:
        if _url_resolves(company["website"]):
            validated.append(company)
        else:
            logger.info(f"[Cold Outreach] Dead URL — skipping {company['name']}")
        time.sleep(1)

    new_companies = validated

    if test_mode:
        new_companies = new_companies[:test_company_limit]
    elif len(new_companies) > max_per_day:
        new_companies = new_companies[:max_per_day]

    if not new_companies:
        print("[Cold Outreach] No new companies found today")
        return []

    print(f"[Cold Outreach] Processing {len(new_companies)} company/companies...")

    # === PARTS 2, 3, 6, 7: Enrich, resolve, draft, create Gmail draft ===
    new_contacts: list = []

    for company in new_companies:
        company_name = company["name"]
        print(f"  {company_name} ({urlparse(company['vc_source']).netloc})")

        # Part 2: Apollo enrichment
        contact = None
        if apollo_key:
            contact = enrich_via_apollo(
                company_name, company["website"], apollo_key,
                min_emp=min_emp, max_emp=max_emp,
            )

        if contact is None:
            # Minimal skeleton when Apollo unavailable or filtered out
            contact = {
                "company_name": company_name,
                "website_url": company["website"],
                "first_name": "",
                "last_name": "",
                "title": "",
                "email": None,
                "email_status": "NOT FOUND",
                "linkedin_url": "",
                "employees": None,
                "domain": "",
            }

        contact["vc_source"] = company.get("vc_source", "")
        contact["date_added"] = str(date.today())

        # Part 3: LinkedIn URL resolution
        li_url, li_type = resolve_linkedin(
            contact.get("first_name", ""),
            contact.get("last_name", ""),
            contact.get("company_name", ""),
            contact.get("linkedin_url", ""),
        )
        contact["linkedin_url"] = li_url
        contact["linkedin_url_type"] = li_type

        # Part 4: Dedup against existing sheet emails
        email_val = (contact.get("email") or "").lower().strip()
        if email_val and email_val in existing_emails:
            logger.info(f"[Cold Outreach] {company_name} already in tracker, skipping")
            continue

        # Part 6: Generate cold email
        draft_text = generate_cold_email(contact, anthropic_key)
        contact["draft_text"] = draft_text
        if draft_text:
            _save_draft_file(company_name, draft_text)

        # Part 7: Create Gmail draft (optional)
        contact["gmail_draft_url"] = ""
        if create_drafts and gmail_service and draft_text:
            subject, body = _parse_subject_and_body(draft_text)
            to_email = contact.get("email", "")
            if to_email and subject and body:
                from_email = os.getenv("YOUR_EMAIL", "")
                _, compose_url = create_gmail_draft(gmail_service, to_email, subject, body, from_email)
                if compose_url:
                    contact["gmail_draft_url"] = compose_url

        new_contacts.append(contact)
        # Track email for in-run dedup
        if email_val:
            existing_emails.add(email_val)

        time.sleep(0.5)

    if not new_contacts:
        print("[Cold Outreach] No new contacts after filtering")
        return []

    # Part 5: Write Excel sheet
    write_cold_outreach_sheet(workbook, new_contacts, sheet_name=sheet_name)

    # Part 4: Persist seen companies (skipped in test/dry-run mode)
    if not test_mode:
        updated_seen = _load_seen_companies()
        for company in new_companies:
            updated_seen.add(company["name"].lower().strip())
        _save_seen_companies(updated_seen)

    drafted_count = sum(1 for c in new_contacts if c.get("draft_text"))
    print(
        f"[Cold Outreach] {len(new_contacts)} new contacts found, "
        f"{drafted_count} emails drafted"
    )
    return new_contacts
