#!/usr/bin/env python3
"""
Cold Outreach Module
Scrapes VC portfolio sites, classifies tech companies, and finds founders/CEOs via
Claude web search.

Standalone:  python cold_outreach.py [--refresh] [--limit N]
Via pipeline: imported by job_pipeline_full.py --cold-outreach
"""

import os
import re
import sys
import json
import time
import argparse
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

load_dotenv(override=True)

# ── Config ────────────────────────────────────────────────────────────────────
CACHE_FILE       = "cold_outreach_cache.json"
CALL_DELAY       = 45         # seconds between Claude API calls (web search ~10k tokens/call, 30k/min limit = 45s safe gap)
RATE_LIMIT_WAIT  = 45         # seconds to wait on rate limit before retry (exponential backoff applied)
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
}

YOUR_PORTFOLIO = os.getenv("YOUR_PORTFOLIO", "")
YOUR_NAME      = os.getenv("YOUR_NAME", "")

# Shared tool-use block counter (importable by job_pipeline_full.py)
tool_use_block_count = 0


# ── Cache helpers ─────────────────────────────────────────────────────────────

def load_cache():
    if Path(CACHE_FILE).exists():
        try:
            with open(CACHE_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


# ── Domain helper ─────────────────────────────────────────────────────────────

def extract_domain(url):
    if not url:
        return ""
    try:
        if not url.startswith("http"):
            url = "https://" + url
        hostname = urlparse(url).hostname or ""
        parts = hostname.split(".")
        strip = {"www", "careers", "jobs", "boards", "apply", "recruiting"}
        while len(parts) > 2 and parts[0] in strip:
            parts = parts[1:]
        return ".".join(parts)
    except Exception:
        return ""


# ── VC Portfolio Scrapers ──────────────────────────────────────────────────────

SKIP_DOMAINS = {
    "linkedin", "twitter", "instagram", "facebook", "youtube", "tiktok",
    "crunchbase", "angel.co", "pitchbook", "mailto", "apple.com",
    "google.com", "github.com", "bloomberg", "techcrunch", "axios",
}

def _is_skip_href(href):
    return any(x in href for x in SKIP_DOMAINS) or href.startswith("#") or not href.startswith("http")

def _dedup(results):
    seen = set()
    out = []
    for r in results:
        k = r["company_name"].lower().strip()
        if k and len(k) > 1 and k not in seen:
            seen.add(k)
            out.append(r)
    return out


def scrape_era_nyc():
    """https://www.eranyc.com/companies/ — server-rendered, article tags per company."""
    results = []
    try:
        r = requests.get("https://www.eranyc.com/companies/", headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")
        for article in soup.select("article"):
            # First heading inside article = company name
            name_el = article.find(["h1", "h2", "h3", "h4", "h5", "strong"])
            if not name_el:
                continue
            name = name_el.get_text(strip=True)
            # External link inside article (not email-protection)
            href = ""
            for a in article.select("a[href^='http']"):
                if "eranyc.com" not in a["href"] and "cdn-cgi" not in a["href"]:
                    href = a["href"]
                    break
            if name and 1 < len(name) < 60:
                results.append({"company_name": name, "website": href, "source": "ERA NYC"})
        deduped = _dedup(results)
        print(f"  [ERA NYC] {len(deduped)} companies found")
        return deduped
    except Exception as e:
        print(f"  [ERA NYC] ERROR: {e}")
        return []


def scrape_fj_labs():
    """https://www.fjlabs.com/portfolio — server-rendered, a[href^=http] links."""
    results = []
    try:
        r = requests.get("https://www.fjlabs.com/portfolio", headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href^='http']"):
            name = a.get_text(strip=True)
            href = a["href"]
            if not name or len(name) < 2 or len(name) > 60:
                continue
            if _is_skip_href(href) or "fjlabs" in href:
                continue
            results.append({"company_name": name, "website": href, "source": "FJ Labs"})
        deduped = _dedup(results)
        print(f"  [FJ Labs] {len(deduped)} companies found")
        return deduped
    except Exception as e:
        print(f"  [FJ Labs] ERROR: {e}")
        return []


def scrape_pioneer():
    """https://pioneer.app/companies — server-rendered YC-style accelerator portfolio."""
    # Skip personal/founder pages and known non-company domains
    personal_skip = {"dcgross.com", "x.com", "twitter.com", "substack.com", "medium.com"}
    results = []
    try:
        r = requests.get("https://pioneer.app/companies", headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href^='http']"):
            name = a.get_text(strip=True)
            href = a["href"]
            if not name or len(name) < 2 or len(name) > 50:
                continue
            if _is_skip_href(href) or "pioneer.app" in href:
                continue
            if any(s in href for s in personal_skip):
                continue
            # Skip names that look like full person names (e.g. "John Smith")
            if re.match(r"^[A-Z][a-z]+ [A-Z][a-z]+$", name):
                continue
            results.append({"company_name": name, "website": href, "source": "Pioneer"})
        deduped = _dedup(results)
        print(f"  [Pioneer] {len(deduped)} companies found")
        return deduped
    except Exception as e:
        print(f"  [Pioneer] ERROR: {e}")
        return []


def scrape_village_global():
    """https://villageglobal.vc/portfolio/ — server-rendered VC portfolio."""
    results = []
    try:
        r = requests.get("https://villageglobal.vc/portfolio/", headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href^='http']"):
            name = a.get_text(strip=True)
            href = a["href"]
            if not name or len(name) < 2 or len(name) > 50:
                continue
            if _is_skip_href(href) or "villageglobal" in href:
                continue
            # Skip names that look like @handles
            if name.startswith("@"):
                continue
            results.append({"company_name": name, "website": href, "source": "Village Global"})
        deduped = _dedup(results)
        print(f"  [Village Global] {len(deduped)} companies found")
        return deduped
    except Exception as e:
        print(f"  [Village Global] ERROR: {e}")
        return []


def scrape_500global():
    """https://500.co/companies — server-rendered list of portfolio companies."""
    results = []
    try:
        r = requests.get("https://500.co/companies", headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.select("a[href^='http']"):
            name = a.get_text(strip=True)
            href = a["href"]
            if not name or len(name) < 2 or len(name) > 60:
                continue
            if _is_skip_href(href) or "500.co" in href or "500startups" in href:
                continue
            results.append({"company_name": name, "website": href, "source": "500 Global"})
        deduped = _dedup(results)
        print(f"  [500 Global] {len(deduped)} companies found")
        return deduped
    except Exception as e:
        print(f"  [500 Global] ERROR: {e}")
        return []


def scrape_betaworks():
    """https://betaworks.com/companies — server-rendered NYC accelerator."""
    results = []
    try:
        r = requests.get("https://betaworks.com/companies", headers=HEADERS, timeout=12)
        soup = BeautifulSoup(r.text, "html.parser")
        # Try headings first (more precise)
        for heading in soup.find_all(["h2", "h3", "h4"]):
            name = heading.get_text(strip=True)
            parent = heading.parent
            href = ""
            if parent:
                link = parent.find("a", href=True)
                if link and link["href"].startswith("http") and "betaworks" not in link["href"]:
                    href = link["href"]
            if name and 1 < len(name) < 60 and not _is_skip_href(href or "http://x"):
                results.append({"company_name": name, "website": href, "source": "Betaworks"})
        # Fallback to all external links if headings yield nothing
        if not results:
            for a in soup.select("a[href^='http']"):
                name = a.get_text(strip=True)
                href = a["href"]
                if not name or len(name) < 2 or len(name) > 60:
                    continue
                if _is_skip_href(href) or "betaworks" in href:
                    continue
                results.append({"company_name": name, "website": href, "source": "Betaworks"})
        deduped = _dedup(results)
        print(f"  [Betaworks] {len(deduped)} companies found")
        return deduped
    except Exception as e:
        print(f"  [Betaworks] ERROR: {e}")
        return []


def scrape_all_sources():
    """Run all scrapers and return combined company list."""
    print("\nScraping VC portfolio sites...")
    all_companies = []
    for fn in [
        scrape_era_nyc,
        scrape_fj_labs,
        scrape_betaworks,
        scrape_pioneer,
        scrape_village_global,
        scrape_500global,
    ]:
        all_companies.extend(fn())
    print(f"\nTotal companies scraped: {len(all_companies)}\n")
    return all_companies


# ── Claude API helpers ────────────────────────────────────────────────────────

def _claude_call(client, prompt, max_tokens=500, retries=3):
    """Call Claude with web search tool. Returns (result_text, tool_use_count)."""
    global tool_use_block_count
    for attempt in range(retries + 1):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=max_tokens,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{"role": "user", "content": prompt}]
            )
            count = sum(
                1 for b in response.content
                if hasattr(b, "type") and b.type in ("tool_use", "server_tool_use")
            )
            tool_use_block_count += count
            text = "".join(
                b.text for b in response.content
                if hasattr(b, "type") and b.type == "text"
            )
            return text, count
        except Exception as e:
            err = str(e)
            if ("rate_limit" in err or "529" in err) and attempt < retries:
                wait = RATE_LIMIT_WAIT * (2 ** attempt)  # 22s, 44s, 88s
                print(f"    [rate limit] waiting {wait}s (attempt {attempt+1}/{retries})...")
                time.sleep(wait)
                continue
            print(f"    [API error] {e}")
            return "", 0
    return "", 0


def _parse_json(text):
    """Strip markdown fences and parse first JSON object or array found."""
    clean = re.sub(r"```json|```", "", text).strip()
    match = re.search(r"(\{.*?\}|\[.*?\])", clean, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except Exception:
            pass
    return None


def get_hq_city(company_name, website, client):
    """Use Claude web search to find the company's headquarters city. Returns a string like
    'Philadelphia, PA', 'Remote', or 'Unknown'."""
    prompt = (
        f"Search the web for the headquarters city of the company '{company_name}' "
        f"(website: {website}). "
        f"Reply with only a short location string such as 'Philadelphia, PA' or 'New York, NY' "
        f"or 'Remote' if the company is fully remote/distributed. "
        f"If you cannot determine it, reply 'Unknown'. "
        f"Reply JSON only: {{\"hq_city\": \"Philadelphia, PA\"}}"
    )
    text, _ = _claude_call(client, prompt, max_tokens=150)
    time.sleep(CALL_DELAY)
    data = _parse_json(text)
    if data and isinstance(data, dict):
        city = data.get("hq_city", "Unknown")
        return city if city else "Unknown"
    return "Unknown"


def is_target_area(city_str):
    """Return True if city_str is in the Philadelphia area, NYC area, remote, or unknown."""
    if not city_str:
        return True
    c = city_str.lower()

    # Always keep if unknown
    if "unknown" in c:
        return True

    # Remote / distributed
    remote_keywords = ["remote", "distributed", "no office", "fully remote"]
    if any(k in c for k in remote_keywords):
        return True

    # Philadelphia area — word-boundary match on 'pa' to avoid false positives
    philly_keywords = [
        "philadelphia", "exton", "king of prussia", "malvern", "wayne",
        "conshohocken", "blue bell", "horsham", "berwyn", "radnor",
        "villanova", "bryn mawr", "west chester", "newtown square",
    ]
    if any(k in c for k in philly_keywords):
        return True
    # standalone 'pa' as a word
    if re.search(r'\bpa\b', c):
        return True

    # NYC area — word-boundary match on 'ny' and 'nyc'
    nyc_keywords = [
        "new york", "manhattan", "brooklyn", "queens", "bronx",
        "jersey city", "hoboken", "newark",
    ]
    if any(k in c for k in nyc_keywords):
        return True
    if re.search(r'\bnyc\b', c) or re.search(r'\bny\b', c):
        return True

    return False


def classify_tech(company_name, website, client):
    """Return True if the company is a software/AI/tech company."""
    prompt = (
        f"Search the web for '{company_name}' (website: {website}). "
        f"Is this a software, AI, or technology company? "
        f"Exclude biotech, pharma, food, apparel, cosmetics, real estate, non-profit. "
        f"Reply JSON only: {{\"is_tech\": true/false, \"category\": \"one word\"}}"
    )
    text, _ = _claude_call(client, prompt, max_tokens=200)
    time.sleep(CALL_DELAY)
    data = _parse_json(text)
    if data and isinstance(data, dict):
        return bool(data.get("is_tech", True))
    return True  # default include if uncertain


def get_founding_year(company_name, website, client):
    """Return founding year as int, or None."""
    prompt = (
        f"Search the web for the founding year of '{company_name}' (website: {website}). "
        f"Reply JSON only: {{\"founded\": 2023}}"
    )
    text, _ = _claude_call(client, prompt, max_tokens=150)
    time.sleep(CALL_DELAY)
    data = _parse_json(text)
    if data and isinstance(data, dict):
        try:
            return int(data.get("founded", 0)) or None
        except Exception:
            pass
    return None


def find_founder(company_name, client):
    """Find CEO or co-founder via Claude web search."""
    global tool_use_block_count
    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1000,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": f"Search LinkedIn and the web right now for the CEO or co-founder of {company_name}. Find their full name, title, LinkedIn URL, and email if publicly available. Return only real people found via web search. Return JSON only: {{\"name\": \"\", \"title\": \"\", \"linkedin_url\": \"\", \"email\": \"\"}}"}]
        )
        count = sum(
            1 for b in response.content
            if hasattr(b, "type") and b.type in ("tool_use", "server_tool_use")
        )
        tool_use_block_count += count
        result_text = ""
        for block in response.content:
            if block.type == "text":
                result_text += block.text
        # Strip ```json fences before parsing
        clean = re.sub(r"```json|```", "", result_text).strip()
        match = re.search(r"(\{.*?\}|\[.*?\])", clean, re.DOTALL)
        if match:
            try:
                data = json.loads(match.group())
                if isinstance(data, dict):
                    return data
                if isinstance(data, list) and data:
                    return data[0]
            except Exception as e:
                print(f"WARNING: [find_founder:{company_name}] JSON parse error: {e}")
        return {"name": "", "title": "", "linkedin_url": "", "email": ""}
    except Exception as e:
        print(f"WARNING: [find_founder:{company_name}] {e}")
        return {"name": "", "title": "", "linkedin_url": "", "email": ""}


# ── Excel output ──────────────────────────────────────────────────────────────

def write_excel(rows):
    """Write results to a timestamped Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    fname = f"cold_outreach_{ts}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Cold Outreach"

    hfill     = PatternFill("solid", fgColor="0F2940")
    new_fill  = PatternFill("solid", fgColor="FFFDE7")  # light yellow for NEW companies
    link_font = Font(name="Arial", color="0563C1", underline="single", size=9)

    headers = [
        "Site", "Company", "Contact Name", "Title", "LinkedIn URL",
        "Contact Date", "Email", "Company Domain",
        "Apollo Lookup", "Email Pattern",
        "Founded", "NEW?"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        is_new  = row.get("is_new", False)
        domain  = row.get("domain", "")
        apollo  = f"https://app.apollo.io/#/people?q_organization_domains[]={domain}" if domain else ""
        linkedin = row.get("linkedin_url", "")

        ws.append([
            row.get("site", ""),
            row.get("company", ""),
            row.get("contact_name", ""),
            row.get("contact_title", ""),
            linkedin,
            row.get("contact_date", ""),
            row.get("email", ""),
            domain,
            "Open Apollo" if apollo else "",
            "",   # Email Pattern — manual
            row.get("founded", ""),
            "NEW" if is_new else "",
        ])

        # Apply yellow fill to NEW rows
        if is_new:
            for cell in ws[row_idx]:
                cell.fill = new_fill

        # Apollo hyperlink (column 9)
        if apollo:
            cell = ws.cell(row=row_idx, column=9)
            cell.hyperlink = apollo
            cell.font      = link_font

        # LinkedIn hyperlink (column 5)
        if linkedin and linkedin.startswith("http"):
            cell = ws.cell(row=row_idx, column=5)
            cell.hyperlink = linkedin
            cell.font      = link_font

    col_widths = [35, 22, 22, 22, 45, 14, 28, 20, 14, 16, 10, 6]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(fname)
    return fname


# ── Pipeline Integration ──────────────────────────────────────────────────────

def run_cold_outreach(client, max_companies=5, refresh=False):
    """
    Scrape VC portfolio sites, research founders, and return result rows.
    Called by job_pipeline_full.py when --cold-outreach is passed.

    Source: VC portfolio scrapers only (ERA NYC, FJ Labs, Betaworks, Pioneer, etc.)
    Never receives or processes companies from the ATS/Jobs pipeline.

    Returns list of dicts with keys:
        site, company, source, contact_name, contact_title,
        linkedin_url, contact_date, email, domain,
        founded, is_new
    """
    cache = {} if refresh else load_cache()

    # Scrape VC portfolio sites — this is the ONLY source for cold outreach
    companies = scrape_all_sources()

    new_companies = [c for c in companies if c["company_name"].lower().strip() not in cache]
    skipped = len(companies) - len(new_companies)
    print(f"Cold Outreach cache: {skipped} already processed, {len(new_companies)} new\n")

    if max_companies:
        new_companies = new_companies[:max_companies]
        print(f"[cold outreach] Researching first {max_companies} new companies\n")

    total = len(new_companies)
    results = []
    today = datetime.now().strftime("%Y-%m-%d")

    for idx, company in enumerate(new_companies, 1):
        name    = company["company_name"].strip()
        website = company.get("website", "")
        source  = company.get("source", "")
        print(f"  Cold outreach [{name}] ({idx}/{total}) -- {source}")

        # 1. Classify as tech or not
        is_tech = classify_tech(name, website, client)
        if not is_tech:
            print(f"    Skipping: not a tech/software company")
            cache[name.lower().strip()] = {"skipped": True, "reason": "non-tech"}
            save_cache(cache)
            continue

        # 2. Geography filter: keep only PHL/NYC area or remote-first companies
        hq_city = get_hq_city(name, website, client)
        if not is_target_area(hq_city):
            print(f"    DROPPED {name} -- {hq_city} (outside target area)")
            cache[name.lower().strip()] = {"skipped": True, "reason": f"outside target area: {hq_city}"}
            save_cache(cache)
            continue
        print(f"    KEPT {name} -- {hq_city}")

        # 3. Get founding year
        founded = get_founding_year(name, website, client)
        is_new  = founded is not None and founded >= 2023

        # 4. Find founder / CEO
        contact       = find_founder(name, client)
        contact_name  = contact.get("name", "")
        contact_title = contact.get("title", "")
        linkedin_url  = contact.get("linkedin_url", "")
        email         = contact.get("email", "")
        print(f"    Contact: {contact_name or '(none found)'} | Email: {email or '(none)'} | Founded: {founded or '?'} {'[NEW]' if is_new else ''}")

        domain = extract_domain(website)
        row = {
            "site":          website,
            "company":       name,
            "source":        source,
            "contact_name":  contact_name,
            "contact_title": contact_title,
            "linkedin_url":  linkedin_url,
            "contact_date":  today,
            "email":         email,
            "domain":        domain,
            "founded":       founded or "",
            "is_new":        is_new,
        }
        results.append(row)

        cache[name.lower().strip()] = {
            "processed_date": today,
            "contact":        contact_name,
            "email":          email,
            "founded":        founded,
        }
        save_cache(cache)

    # Sort: NEW companies (2023+) first, then alphabetical
    results.sort(key=lambda r: (0 if r.get("is_new") else 1, r["company"]))
    return results


# ── Main (standalone) ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh", action="store_true", help="Reprocess all companies, ignoring cache")
    parser.add_argument("--limit",   type=int, default=None, help="Max companies to process per run")
    args = parser.parse_args()

    import anthropic
    api_key = os.getenv("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        print("Set ANTHROPIC_API_KEY in .env")
        sys.exit(1)
    client = anthropic.Anthropic(api_key=api_key)

    # Load cache
    cache = {} if args.refresh else load_cache()
    if args.refresh:
        print("[--refresh] Ignoring cache, reprocessing all companies.\n")

    # Scrape all VC portfolio sources
    companies = scrape_all_sources()

    # Filter out cached companies
    new_companies = [c for c in companies if c["company_name"].lower().strip() not in cache]
    skipped = len(companies) - len(new_companies)
    print(f"Cache: {skipped} companies skipped (already processed), {len(new_companies)} new\n")

    if args.limit:
        new_companies = new_companies[:args.limit]
        print(f"[--limit] Processing first {args.limit} companies\n")

    total = len(new_companies)
    results = []
    today   = datetime.now().strftime("%Y-%m-%d")

    for idx, company in enumerate(new_companies, 1):
        name    = company["company_name"].strip()
        website = company.get("website", "")
        source  = company.get("source", "")
        print(f"Researching [{name}] ({idx}/{total}) -- {source}")

        # 1. Classify as tech or not
        is_tech = classify_tech(name, website, client)
        if not is_tech:
            print(f"  Skipping: not a tech/software company")
            cache[name.lower().strip()] = {"skipped": True, "reason": "non-tech"}
            save_cache(cache)
            continue

        # 2. Geography filter: keep only PHL/NYC area or remote-first companies
        hq_city = get_hq_city(name, website, client)
        if not is_target_area(hq_city):
            print(f"  DROPPED {name} -- {hq_city} (outside target area)")
            cache[name.lower().strip()] = {"skipped": True, "reason": f"outside target area: {hq_city}"}
            save_cache(cache)
            continue
        print(f"  KEPT {name} -- {hq_city}")

        # 3. Get founding year
        founded = get_founding_year(name, website, client)
        is_new  = founded is not None and founded >= 2023

        # 4. Find founder / CEO
        contact = find_founder(name, client)
        contact_name  = contact.get("name", "")
        contact_title = contact.get("title", "")
        linkedin_url  = contact.get("linkedin_url", "")
        email         = contact.get("email", "")
        print(f"  Contact: {contact_name or '(none found)'} | Email: {email or '(none)'} | Founded: {founded or '?'} {'[NEW]' if is_new else ''}")

        domain = extract_domain(website)
        row = {
            "site":          website,
            "company":       name,
            "source":        source,
            "contact_name":  contact_name,
            "contact_title": contact_title,
            "linkedin_url":  linkedin_url,
            "contact_date":  today,
            "email":         email,
            "domain":        domain,
            "founded":       founded or "",
            "is_new":        is_new,
        }
        results.append(row)

        # Update cache
        cache[name.lower().strip()] = {
            "processed_date": today,
            "contact":        contact_name,
            "email":          email,
            "founded":        founded,
        }
        save_cache(cache)

    # Sort: NEW companies first
    results.sort(key=lambda r: (0 if r.get("is_new") else 1, r["company"]))

    # Write Excel
    fname = ""
    if results:
        fname = write_excel(results)
    else:
        print("\nNo new companies to process.")

    # Summary
    contacts_found = sum(1 for r in results if r.get("contact_name"))
    emails_found   = sum(1 for r in results if r.get("email"))
    new_count      = sum(1 for r in results if r.get("is_new"))

    print(f"\n{'='*60}")
    print(f"  COLD OUTREACH COMPLETE")
    print(f"  Total scraped:          {len(companies)}")
    print(f"  Skipped (cache):        {skipped}")
    print(f"  Processed this run:     {len(results)}")
    print(f"  NEW companies (2023+):  {new_count}")
    print(f"  Contacts found:         {contacts_found}")
    print(f"  Emails found:           {emails_found}")
    print(f"  Tool-use blocks:        {tool_use_block_count}")
    if tool_use_block_count == 0:
        print("  WARNING: 0 tool_use blocks -- web search may not be invoking correctly")
    else:
        print(f"  OK: web search confirmed active ({tool_use_block_count} search blocks)")
    if fname:
        print(f"  File: {fname}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
